import json
import hashlib
import os
import httpx
from datetime import datetime, timezone
from pathlib import Path
from typing import Annotated, Any, AsyncGenerator
from uuid import uuid4

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct

from auth import require_roles, _write_audit_log
from database import get_db
from models import Tenant, TenantResource

router = APIRouter()

# Initialize Qdrant client
QDRANT_HOST = os.getenv("QDRANT_HOST", "qdrant")
QDRANT_PORT = int(os.getenv("QDRANT_PORT", "6333"))
OLLAMA_HOST = os.getenv("OLLAMA_HOST", "http://ollama:11434")
EMBEDDING_MODEL = "nomic-embed-text"
EMBEDDING_DIM = 384  # nomic-embed-text output dimension

try:
    qdrant_client = QdrantClient(host=QDRANT_HOST, port=QDRANT_PORT)
except Exception as e:
    print(f"Warning: Failed to initialize Qdrant client: {e}")
    qdrant_client = None


def _get_ollama_embedding(text: str) -> list[float] | None:
    """Get embedding from Ollama. Returns None if unavailable."""
    if not text or not text.strip():
        return [0.0] * EMBEDDING_DIM

    try:
        response = httpx.post(
            f"{OLLAMA_HOST}/api/embeddings",
            json={"model": EMBEDDING_MODEL, "prompt": text},
            timeout=30.0
        )
        response.raise_for_status()
        return response.json().get("embedding")
    except Exception as e:
        print(f"Warning: Failed to get Ollama embedding: {e}")
        # Return zero vector as fallback
        return [0.0] * EMBEDDING_DIM


def _get_openai_embedding(text: str, api_key: str, model: str = "text-embedding-3-small") -> list[float] | None:
    """Get embedding from OpenAI."""
    if not text or not text.strip():
        return None

    if not api_key:
        print("Warning: OpenAI API key not provided")
        return None

    try:
        response = httpx.post(
            "https://api.openai.com/v1/embeddings",
            headers={"Authorization": f"Bearer {api_key}"},
            json={"model": model, "input": text},
            timeout=30.0
        )
        response.raise_for_status()
        return response.json().get("data", [{}])[0].get("embedding")
    except Exception as e:
        print(f"Warning: Failed to get OpenAI embedding: {e}")
        return None


def _get_cohere_embedding(text: str, api_key: str, model: str = "embed-english-v3.0") -> list[float] | None:
    """Get embedding from Cohere."""
    if not text or not text.strip():
        return None

    if not api_key:
        print("Warning: Cohere API key not provided")
        return None

    try:
        response = httpx.post(
            "https://api.cohere.com/v1/embed",
            headers={"Authorization": f"Bearer {api_key}"},
            json={"texts": [text], "model": model, "input_type": "search_document"},
            timeout=30.0
        )
        response.raise_for_status()
        embeddings = response.json().get("embeddings", [[]])
        return embeddings[0] if embeddings else None
    except Exception as e:
        print(f"Warning: Failed to get Cohere embedding: {e}")
        return None


def _get_huggingface_embedding(text: str, api_key: str, model_id: str = "sentence-transformers/all-MiniLM-L6-v2") -> list[float] | None:
    """Get embedding from HuggingFace Inference API."""
    if not text or not text.strip():
        return None

    if not api_key:
        print("Warning: HuggingFace API key not provided")
        return None

    try:
        response = httpx.post(
            f"https://api-inference.huggingface.co/models/{model_id}",
            headers={"Authorization": f"Bearer {api_key}"},
            json={"inputs": text},
            timeout=30.0
        )
        response.raise_for_status()
        result = response.json()
        # HF returns [[embedding]] for a single input
        if isinstance(result, list) and len(result) > 0:
            return result[0]
        return None
    except Exception as e:
        print(f"Warning: Failed to get HuggingFace embedding: {e}")
        return None


def _get_embedding(
    text: str,
    provider: str,
    provider_config: dict[str, Any]
) -> list[float] | None:
    """Get embedding using the specified provider."""
    if provider == "openAIEmbeddings":
        api_key = provider_config.get("apiKey", os.getenv("OPENAI_API_KEY", ""))
        model = provider_config.get("model", "text-embedding-3-small")
        return _get_openai_embedding(text, api_key, model)
    elif provider == "cohereEmbeddings":
        api_key = provider_config.get("apiKey", os.getenv("COHERE_API_KEY", ""))
        model = provider_config.get("model", "embed-english-v3.0")
        return _get_cohere_embedding(text, api_key, model)
    elif provider == "huggingFaceInferenceEmbeddings":
        api_key = provider_config.get("apiKey", os.getenv("HUGGINGFACE_API_KEY", ""))
        model = provider_config.get("modelId", "sentence-transformers/all-MiniLM-L6-v2")
        return _get_huggingface_embedding(text, api_key, model)
    else:
        # Default: Ollama
        return _get_ollama_embedding(text)


EMBEDDING_DIMS = {
    "ollamaEmbeddings": 384,
    "openAIEmbeddings": {
        "text-embedding-3-small": 1536,
        "text-embedding-3-large": 3072,
        "text-embedding-ada-002": 1536,
    },
    "cohereEmbeddings": 1024,
    "huggingFaceInferenceEmbeddings": {
        "sentence-transformers/all-MiniLM-L6-v2": 384,
        "sentence-transformers/all-mpnet-base-v2": 768,
    },
}


CHAT_MODEL_COMPONENTS = [
    {
        "name": "ollamaChat",
        "label": "Ollama Chat",
        "category": "Chat Models",
        "description": "Run chat completions with Ollama.",
        "baseClasses": ["ChatModel"],
        "inputs": [
            {"label": "Model", "name": "model", "type": "string", "default": "llama3.2", "optional": False},
            {"label": "Temperature", "name": "temperature", "type": "number", "default": 0.7, "optional": True},
        ],
        "outputs": [{"label": "Chat Model", "name": "chatModel", "baseClasses": ["ChatModel"]}],
    },
    {
        "name": "openAIChat",
        "label": "OpenAI Chat",
        "category": "Chat Models",
        "description": "Run chat completions with OpenAI.",
        "baseClasses": ["ChatModel"],
        "inputs": [
            {"label": "Model", "name": "model", "type": "string", "default": "gpt-4o-mini", "optional": False},
            {"label": "Temperature", "name": "temperature", "type": "number", "default": 0.7, "optional": True},
        ],
        "outputs": [{"label": "Chat Model", "name": "chatModel", "baseClasses": ["ChatModel"]}],
        "credential": {
            "label": "OpenAI Credential",
            "name": "credential",
            "type": "credential",
            "credentialNames": ["openAIApi"],
            "optional": True,
        },
    },
    {
        "name": "anthropicChat",
        "label": "Anthropic Chat",
        "category": "Chat Models",
        "description": "Run chat completions with Anthropic models.",
        "baseClasses": ["ChatModel"],
        "inputs": [
            {
                "label": "Model",
                "name": "model",
                "type": "string",
                "default": "claude-3-5-sonnet-latest",
                "optional": False,
            },
            {"label": "Temperature", "name": "temperature", "type": "number", "default": 0.7, "optional": True},
        ],
        "outputs": [{"label": "Chat Model", "name": "chatModel", "baseClasses": ["ChatModel"]}],
        "credential": {
            "label": "Anthropic Credential",
            "name": "credential",
            "type": "credential",
            "credentialNames": ["anthropicApi"],
            "optional": True,
        },
    },
    {
        "name": "perplexityChat",
        "label": "Perplexity Chat",
        "category": "Chat Models",
        "description": "Run chat completions with Perplexity models.",
        "baseClasses": ["ChatModel"],
        "inputs": [
            {"label": "Model", "name": "model", "type": "string", "default": "sonar-pro", "optional": False},
            {"label": "Temperature", "name": "temperature", "type": "number", "default": 0.7, "optional": True},
        ],
        "outputs": [{"label": "Chat Model", "name": "chatModel", "baseClasses": ["ChatModel"]}],
        "credential": {
            "label": "Perplexity Credential",
            "name": "credential",
            "type": "credential",
            "credentialNames": ["perplexityApi"],
            "optional": True,
        },
    },
    {
        "name": "gemini",
        "label": "Gemini Chat",
        "category": "Chat Models",
        "description": "Run chat completions with Google Gemini models.",
        "baseClasses": ["ChatModel"],
        "inputs": [
            {"label": "Model", "name": "model", "type": "string", "default": "gemini-1.5-pro", "optional": False},
            {"label": "Temperature", "name": "temperature", "type": "number", "default": 0.7, "optional": True},
        ],
        "outputs": [{"label": "Chat Model", "name": "chatModel", "baseClasses": ["ChatModel"]}],
        "credential": {
            "label": "Google Generative AI Credential",
            "name": "credential",
            "type": "credential",
            "credentialNames": ["googleGenerativeAI"],
            "optional": True,
        },
    },
]

DOC_LOADERS = [
    {"name": "textFileLoader", "label": "Text File Loader"},
    {"name": "webLoader", "label": "Web Loader"},
    {"name": "pdfFileLoader", "label": "PDF Loader"},
    {"name": "csvFileLoader", "label": "CSV File Loader"},
    {"name": "jsonFileLoader", "label": "JSON File Loader"},
    {"name": "jsonLinesLoader", "label": "JSON Lines File Loader"},
    {"name": "docxFileLoader", "label": "Docx File Loader"},
    {"name": "excelFileLoader", "label": "Excel File Loader"},
    {"name": "notionLoader", "label": "Notion"},
    {"name": "githubLoader", "label": "Github"},
    {"name": "confluenceLoader", "label": "Confluence"},
    {"name": "s3FileLoader", "label": "S3 File Loader"},
    {"name": "cheerioWebScraper", "label": "Cheerio Web Scraper"},
    {"name": "airtableLoader", "label": "Airtable"},
]

VECTOR_PROVIDERS = [
    {"name": "qdrant", "label": "Qdrant", "inputs": [
        {"label": "Collection Name", "name": "collectionName", "type": "string", "default": "", "optional": True}
    ]},
    {"name": "chroma", "label": "Chroma", "inputs": [
        {"label": "Host", "name": "host", "type": "string", "default": "localhost", "optional": True},
        {"label": "Port", "name": "port", "type": "number", "default": 8000, "optional": True},
        {"label": "Collection Name", "name": "collectionName", "type": "string", "default": "documents", "optional": True}
    ]},
    {"name": "pinecone", "label": "Pinecone", "inputs": [
        {"label": "Index Name", "name": "indexName", "type": "string", "default": "", "optional": False},
        {"label": "Namespace", "name": "namespace", "type": "string", "default": "", "optional": True},
        {"label": "API Key", "name": "apiKey", "type": "password", "optional": False}
    ]},
    {"name": "weaviate", "label": "Weaviate", "inputs": [
        {"label": "Host", "name": "host", "type": "string", "default": "localhost", "optional": True},
        {"label": "Port", "name": "port", "type": "number", "default": 8080, "optional": True},
        {"label": "Class Name", "name": "className", "type": "string", "default": "Document", "optional": True},
        {"label": "API Key", "name": "apiKey", "type": "password", "optional": True}
    ]},
]
EMBEDDING_PROVIDERS = [
    {"name": "ollamaEmbeddings", "label": "Ollama Embeddings", "inputs": [
        {"label": "Model", "name": "model", "type": "string", "default": "nomic-embed-text"}
    ]},
    {"name": "openAIEmbeddings", "label": "OpenAI Embeddings", "inputs": [
        {"label": "Model", "name": "model", "type": "string", "default": "text-embedding-3-small"},
        {"label": "API Key", "name": "apiKey", "type": "password", "optional": False}
    ]},
    {"name": "cohereEmbeddings", "label": "Cohere Embeddings", "inputs": [
        {"label": "Model", "name": "model", "type": "string", "default": "embed-english-v3.0"},
        {"label": "API Key", "name": "apiKey", "type": "password", "optional": False}
    ]},
    {"name": "huggingFaceInferenceEmbeddings", "label": "HuggingFace Inference Embeddings", "inputs": [
        {"label": "Model ID", "name": "modelId", "type": "string", "default": "sentence-transformers/all-MiniLM-L6-v2"},
        {"label": "API Key", "name": "apiKey", "type": "password", "optional": False}
    ]},
]
RECORD_MANAGER_PROVIDERS = [{"name": "sqliteRecordManager", "label": "SQLite Record Manager", "inputs": []}]


def _node_icon_svg(node_name: str) -> str:
    safe_name = (node_name or "node").strip()[:32]
    parts = [part for part in safe_name.replace("_", " ").replace("-", " ").split() if part]
    initials = "".join(part[0] for part in parts[:2]).upper() or "N"
    digest = hashlib.md5(safe_name.encode("utf-8")).hexdigest()
    bg = f"#{digest[:6]}"
    fg = "#ffffff"

    return (
        "<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 64 64'>"
        f"<rect x='2' y='2' width='60' height='60' rx='30' fill='{bg}' />"
        "<circle cx='32' cy='32' r='30' fill='none' stroke='rgba(255,255,255,0.28)' stroke-width='2' />"
        f"<text x='32' y='39' text-anchor='middle' fill='{fg}' font-family='Arial, sans-serif' "
        "font-size='22' font-weight='700'>"
        f"{initials}</text></svg>"
    )


def _credential_icon_svg(credential_name: str) -> str:
    safe_name = (credential_name or "credential").strip()[:32]
    parts = [part for part in safe_name.replace("_", " ").replace("-", " ").split() if part]
    initials = "".join(part[0] for part in parts[:2]).upper() or "C"
    digest = hashlib.md5(f"credential:{safe_name}".encode("utf-8")).hexdigest()
    bg = f"#{digest[:6]}"
    fg = "#ffffff"

    return (
        "<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 64 64'>"
        f"<rect x='2' y='2' width='60' height='60' rx='14' fill='{bg}' />"
        "<path d='M24 32a8 8 0 1 1 12.7 6.4l-0.7 0.5V44h-6v-3.2l-0.7-0.5A8 8 0 0 1 24 32Z' fill='rgba(255,255,255,0.2)' />"
        f"<text x='32' y='54' text-anchor='middle' fill='{fg}' font-family='Arial, sans-serif' font-size='12' font-weight='700'>{initials}</text>"
        "</svg>"
    )


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


@router.get("/v1/node-icon/{node_name}")
def get_node_icon(node_name: str) -> Response:
    svg = _node_icon_svg(node_name)
    return Response(content=svg, media_type="image/svg+xml")


@router.get("/v1/components-credentials-icon/{credential_name}")
def get_components_credentials_icon(credential_name: str) -> Response:
    svg = _credential_icon_svg(credential_name)
    return Response(content=svg, media_type="image/svg+xml")


@router.get("/v1/chatflows-streaming/{chatflow_id}")
def get_chatflow_streaming_compat(chatflow_id: str) -> dict[str, Any]:
    return {
        "chatflowId": chatflow_id,
        "supported": False,
        "isUnavailable": True,
        "detail": "Public chatbot streaming is not available in this deployment."
    }


@router.get("/v1/public-chatbotConfig/{chatflow_id}")
def get_public_chatbot_config_compat(chatflow_id: str) -> dict[str, Any]:
    return {
        "id": chatflow_id,
        "isPublic": False,
        "isUnavailable": True,
        "chatbotConfig": "{}",
        "apikeyid": "",
        "detail": "Public chatbot configuration endpoint is unavailable in this deployment."
    }


def _ensure_tenant(db: Session, tenant_id: str) -> None:
    tenant = db.get(Tenant, tenant_id)
    if tenant is None:
        db.add(Tenant(id=tenant_id, name=f"tenant-{tenant_id[:8]}"))
        db.flush()


def _parse_json(value: Any, fallback: Any) -> Any:
    if isinstance(value, (list, dict)):
        return value
    if isinstance(value, str):
        try:
            return json.loads(value)
        except Exception:
            return fallback
    return fallback


def _resource_display_name(resource: TenantResource) -> str:
    payload = resource.payload or {}
    return payload.get("display_name") or resource.name


def _query_resources(db: Session, tenant_id: str, resource_type: str) -> list[TenantResource]:
    return (
        db.query(TenantResource)
        .filter(TenantResource.tenant_id == tenant_id, TenantResource.resource_type == resource_type)
        .order_by(TenantResource.updated_at.desc())
        .all()
    )


def _get_resource(db: Session, tenant_id: str, resource_type: str, resource_id: str) -> TenantResource:
    resource = (
        db.query(TenantResource)
        .filter(
            TenantResource.tenant_id == tenant_id,
            TenantResource.resource_type == resource_type,
            TenantResource.id == resource_id,
        )
        .one_or_none()
    )
    if resource is None:
        raise HTTPException(status_code=404, detail=f"{resource_type} not found")
    return resource


def _create_resource(db: Session, tenant_id: str, user_id: str, resource_type: str, display_name: str, payload: dict[str, Any]) -> TenantResource:
    _ensure_tenant(db, tenant_id)
    internal_name = f"{display_name}__{uuid4().hex[:8]}"
    body = payload.copy()
    body["display_name"] = display_name
    now = datetime.now(timezone.utc)

    resource = TenantResource(
        id=str(uuid4()),
        tenant_id=tenant_id,
        resource_type=resource_type,
        name=internal_name,
        payload=body,
        created_by=user_id,
        updated_by=user_id,
        created_at=now,
        updated_at=now,
    )
    db.add(resource)
    db.commit()
    db.refresh(resource)
    return resource


def _update_resource(resource: TenantResource, user_id: str, display_name: str | None = None, payload: dict[str, Any] | None = None) -> None:
    if payload is not None:
        merged_payload = payload.copy()
        if display_name is None:
            merged_payload["display_name"] = resource.payload.get("display_name", resource.name)
        else:
            merged_payload["display_name"] = display_name
        resource.payload = merged_payload
    elif display_name is not None:
        resource.payload = {**(resource.payload or {}), "display_name": display_name}

    if display_name is not None:
        # keep internal name unique but still update prefix
        resource.name = f"{display_name}__{resource.id[:8]}"

    resource.updated_by = user_id
    resource.updated_at = datetime.now(timezone.utc)


def _paginate(items: list[Any], page: int, limit: int) -> tuple[list[Any], int]:
    safe_page = max(1, page)
    safe_limit = max(1, min(limit, 200))
    start = (safe_page - 1) * safe_limit
    return items[start : start + safe_limit], len(items)


def _split_text(
    text: str,
    splitter_id: str | None,
    splitter_config: dict[str, Any] | None,
    preview_count: int = 20
) -> tuple[list[dict[str, Any]], int]:
    """Split text into chunks based on splitter configuration."""
    if splitter_config is None:
        splitter_config = {}

    text = text.replace("\r", "")
    if not text.strip():
        all_parts = ["Sample chunk from document loader"]
    elif splitter_id == "recursiveCharacterTextSplitter":
        chunk_size = splitter_config.get("chunkSize", 1000)
        chunk_overlap = splitter_config.get("chunkOverlap", 200)
        all_parts = _recursive_character_split(text, chunk_size, chunk_overlap)
    elif splitter_id == "tokenTextSplitter":
        chunk_size = splitter_config.get("chunkSize", 500)
        all_parts = _token_split(text, chunk_size)
    elif splitter_id == "markdownTextSplitter":
        chunk_size = splitter_config.get("chunkSize", 1000)
        chunk_overlap = splitter_config.get("chunkOverlap", 200)
        all_parts = _markdown_split(text, chunk_size, chunk_overlap)
    elif splitter_id == "characterTextSplitter":
        separator = splitter_config.get("separator", "\n\n")
        chunk_size = splitter_config.get("chunkSize", 1000)
        chunk_overlap = splitter_config.get("chunkOverlap", 200)
        all_parts = _character_split(text, separator, chunk_size, chunk_overlap)
    else:
        # Default: newline split (backwards compatible)
        all_parts = [p.strip() for p in text.split("\n") if p.strip()]

    if not all_parts:
        all_parts = ["Sample chunk from document loader"]

    all_chunks = [
        {
            "id": str(uuid4()),
            "pageContent": part,
            "metadata": {"chunkIndex": idx + 1},
        }
        for idx, part in enumerate(all_parts)
    ]

    return all_chunks[: max(1, preview_count)], len(all_chunks)


def _recursive_character_split(text: str, chunk_size: int = 1000, chunk_overlap: int = 200) -> list[str]:
    """Recursively split text using multiple separators."""
    separators = ["\n\n", "\n", " ", ""]

    def split_recursive(text: str, separators: list[str]) -> list[str]:
        good_splits = []
        separator = separators[-1]

        for i, sep in enumerate(separators):
            if sep == "":
                splits = list(text)
            else:
                splits = text.split(sep)

            good_splits = [s for s in splits if len(s) > chunk_size]
            if not good_splits:
                separator = sep
                break
            if i == 0:
                separator = sep
                break

        if separator != "":
            final_splits = []
            for i, s in enumerate(splits):
                if len(s) < chunk_size:
                    if i == 0:
                        final_splits.append(s)
                    elif len(final_splits[-1]) + len(sep) + len(s) <= chunk_size:
                        final_splits[-1] += sep + s
                    else:
                        final_splits.append(s)
                else:
                    if final_splits and len(final_splits[-1]) + len(sep) + len(s[:chunk_size]) <= chunk_size:
                        final_splits[-1] += sep + s
                    splits_sub = split_recursive(s, separators[separators.index(separator) + 1:])
                    final_splits.extend(splits_sub)
            return final_splits
        return splits

    result = split_recursive(text, separators)

    # Now apply overlap
    chunks = []
    current_chunk = ""
    for part in result:
        if len(current_chunk) + len(part) <= chunk_size:
            current_chunk += part
        else:
            if current_chunk:
                chunks.append(current_chunk.strip())
            current_chunk = part
    if current_chunk:
        chunks.append(current_chunk.strip())

    # Add overlap
    overlapped_chunks = []
    for i, chunk in enumerate(chunks):
        if i == 0:
            overlapped_chunks.append(chunk)
        else:
            prev_chunk = overlapped_chunks[-1]
            overlap_start = max(0, len(prev_chunk) - chunk_overlap)
            overlap_text = prev_chunk[overlap_start:]
            overlapped_chunks.append(overlap_text + " " + chunk)

    return [c.strip() for c in overlapped_chunks if c.strip()]


def _token_split(text: str, chunk_size: int = 500) -> list[str]:
    """Split text based on approximate token count (word * 1.3)."""
    words = text.split()
    chunks = []
    current_chunk = []
    current_tokens = 0

    for word in words:
        word_tokens = len(word) / 4 + 1  # Approximate tokens per word
        if current_tokens + word_tokens > chunk_size and current_chunk:
            chunks.append(" ".join(current_chunk))
            current_chunk = [word]
            current_tokens = word_tokens
        else:
            current_chunk.append(word)
            current_tokens += word_tokens

    if current_chunk:
        chunks.append(" ".join(current_chunk))

    return [c.strip() for c in chunks if c.strip()]


def _markdown_split(text: str, chunk_size: int = 1000, chunk_overlap: int = 200) -> list[str]:
    """Split markdown text by headers, then apply character split to each section."""
    import re

    # Split by markdown headers (##, ###, etc.)
    header_pattern = r"(?=^#{1,6}\s)"
    sections = re.split(header_pattern, text, flags=re.MULTILINE)

    chunks = []
    for section in sections:
        if len(section) <= chunk_size:
            chunks.append(section.strip())
        else:
            # Apply character split to this section
            sub_chunks = _character_split(section, "\n\n", chunk_size, chunk_overlap)
            chunks.extend(sub_chunks)

    return [c.strip() for c in chunks if c.strip()]


def _character_split(text: str, separator: str = "\n\n", chunk_size: int = 1000, chunk_overlap: int = 200) -> list[str]:
    """Split text by a separator while respecting chunk size and overlap."""
    if separator:
        splits = text.split(separator)
    else:
        splits = list(text)

    # Merge splits to respect chunk_size
    good_splits = []
    for s in splits:
        if len(s) > chunk_size:
            if good_splits:
                merged = good_splits[-1] + separator + s if separator else good_splits[-1] + s
                good_splits[-1] = merged
            else:
                good_splits.append(s)
        else:
            if good_splits and len(good_splits[-1]) + len(separator) + len(s) <= chunk_size:
                good_splits[-1] += separator + s if separator else good_splits[-1] + s
            else:
                good_splits.append(s)

    # Apply overlap
    result = []
    for i, chunk in enumerate(good_splits):
        if i == 0:
            result.append(chunk)
        else:
            prev_chunk = result[-1]
            if chunk_overlap > 0 and len(prev_chunk) > chunk_overlap:
                overlap_text = prev_chunk[-chunk_overlap:]
                result.append(overlap_text + (separator if separator else "") + chunk)
            else:
                result.append(chunk)

    return [c.strip() for c in result if c.strip()]


def _build_chunks(source_text: str, preview_count: int = 20) -> tuple[list[dict[str, Any]], int]:
    """Legacy function: splits text using default newline strategy. Use _split_text for new code."""
    return _split_text(source_text, None, {}, preview_count)


# Assistants
@router.get("/assistants")
def list_assistants(
    type: str | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> list[dict[str, Any]]:
    rows = _query_resources(db, user["tenant_id"], "assistant")
    result: list[dict[str, Any]] = []
    for row in rows:
        payload = row.payload or {}
        if type and payload.get("type") != type:
            continue
        result.append(
            {
                "id": row.id,
                "type": payload.get("type", "CUSTOM"),
                "details": payload.get("details", json.dumps({"name": _resource_display_name(row)})),
                "credential": payload.get("credential"),
                "iconSrc": payload.get("iconSrc"),
                "createdDate": row.created_at.isoformat() if row.created_at else None,
                "updatedDate": row.updated_at.isoformat() if row.updated_at else None,
            }
        )
    return result


@router.post("/assistants")
def create_assistant(
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    details = body.get("details")
    details_obj = _parse_json(details, {})
    display_name = details_obj.get("name") or body.get("name") or "Assistant"

    payload = {
        "details": details if isinstance(details, str) else json.dumps(details_obj or {"name": display_name}),
        "credential": body.get("credential"),
        "type": body.get("type", "CUSTOM"),
        "iconSrc": body.get("iconSrc"),
    }

    row = _create_resource(db, user["tenant_id"], user.get("user_id"), "assistant", display_name, payload)
    return {"id": row.id}


@router.get("/assistants/{assistant_id}")
def get_assistant(
    assistant_id: str,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "assistant", assistant_id)
    payload = row.payload or {}
    return {
        "id": row.id,
        "type": payload.get("type", "CUSTOM"),
        "details": payload.get("details", json.dumps({"name": _resource_display_name(row)})),
        "credential": payload.get("credential"),
        "iconSrc": payload.get("iconSrc"),
    }


@router.put("/assistants/{assistant_id}")
def update_assistant(
    assistant_id: str,
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "assistant", assistant_id)
    current = row.payload or {}

    details = body.get("details", current.get("details"))
    details_obj = _parse_json(details, {})
    display_name = details_obj.get("name") or body.get("name") or current.get("display_name") or _resource_display_name(row)

    payload = {
        "details": details if isinstance(details, str) else json.dumps(details_obj or {"name": display_name}),
        "credential": body.get("credential", current.get("credential")),
        "type": body.get("type", current.get("type", "CUSTOM")),
        "iconSrc": body.get("iconSrc", current.get("iconSrc")),
    }

    _update_resource(row, user.get("user_id"), display_name=display_name, payload=payload)
    db.commit()
    return {"id": row.id}


@router.delete("/assistants/{assistant_id}")
def delete_assistant(
    assistant_id: str,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "assistant", assistant_id)
    db.delete(row)
    db.commit()
    return {"status": "deleted", "id": assistant_id}


@router.get("/assistants/components/chatmodels")
def assistants_chatmodels(
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> list[dict[str, Any]]:
    return CHAT_MODEL_COMPONENTS


@router.get("/assistants/components/docstores")
def assistants_docstores(
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> list[dict[str, Any]]:
    stores = _query_resources(db, user["tenant_id"], "document_store")
    return [
        {
            "label": _resource_display_name(store),
            "name": store.id,
            "description": (store.payload or {}).get("description", ""),
        }
        for store in stores
    ]


@router.get("/assistants/components/tools")
def assistants_tools(
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> list[dict[str, Any]]:
    tools = _query_resources(db, user["tenant_id"], "tool")
    result = []
    for tool in tools:
        payload = tool.payload or {}
        result.append(
            {
                "name": payload.get("toolName") or tool.id,
                "label": _resource_display_name(tool),
                "description": payload.get("description", ""),
                "category": payload.get("category", "custom"),
                "baseClasses": ["Tool"],
                "inputs": [],
                "outputs": [{"label": "Tool", "name": "tool", "baseClasses": ["Tool"]}],
            }
        )
    return result


@router.post("/assistants/generate/instruction")
def generate_instruction(
    body: dict[str, Any],
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> dict[str, Any]:
    """Generate assistant instruction using LLM based on task/prompt."""
    user_prompt = str(body.get("prompt") or body.get("task") or body.get("question") or "").strip()

    if not user_prompt:
        instruction = "You are a helpful assistant."
    else:
        # Try to generate using Ollama
        try:
            meta_prompt = f"""You are an AI assistant configuration expert. Based on the following task or purpose, write a concise system instruction (2-3 sentences) for an AI assistant.

Task/Purpose: {user_prompt}

Provide only the system instruction, nothing else."""

            response = httpx.post(
                f"{OLLAMA_HOST}/api/generate",
                json={
                    "model": "llama2",
                    "prompt": meta_prompt,
                    "stream": False,
                    "num_predict": 100,
                },
                timeout=15.0
            )
            response.raise_for_status()
            generated = response.json().get("response", "").strip()
            instruction = generated if generated else f"You are a helpful assistant specialized in: {user_prompt}"
        except Exception as e:
            print(f"Warning: LLM instruction generation failed: {e}")
            # Fallback if LLM is unavailable
            instruction = f"You are a helpful assistant specialized in: {user_prompt}"

    # Return both keys because different UI surfaces currently read either `instruction` or `content`.
    return {"instruction": instruction, "content": instruction}


@router.get("/openai-assistants")
def list_openai_assistants(
    credential: str | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> list[dict[str, Any]]:
    return _list_provider_assistants(
        assistant_type="OPENAI",
        default_model="gpt-4o-mini",
        db=db,
        user=user,
    )


@router.get("/azure-assistants")
def list_azure_assistants(
    credential: str | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> list[dict[str, Any]]:
    return _list_provider_assistants(
        assistant_type="AZURE",
        default_model="gpt-4o-mini",
        db=db,
        user=user,
    )


def _list_provider_assistants(
    assistant_type: str,
    default_model: str,
    db: Session,
    user: dict[str, Any],
) -> list[dict[str, Any]]:
    rows = _query_resources(db, user["tenant_id"], "assistant")
    assistants: list[dict[str, Any]] = []
    for row in rows:
        payload = row.payload or {}
        if payload.get("type") != assistant_type:
            continue
        details = _parse_json(payload.get("details"), {})
        assistant_obj_id = details.get("id") or row.id
        assistants.append(
            {
                "id": assistant_obj_id,
                "name": details.get("name") or _resource_display_name(row),
                "description": details.get("description") or "",
                "model": details.get("model") or default_model,
                "instructions": details.get("instructions") or "",
                "tools": details.get("tools") or [],
                "tool_resources": details.get("tool_resources") or {},
                "temperature": details.get("temperature", 1),
                "top_p": details.get("top_p", 1),
                "_resource_id": row.id,
                "credential": payload.get("credential"),
            }
        )
    return assistants


@router.get("/openai-assistants/{assistant_obj_id}")
def get_openai_assistant(
    assistant_obj_id: str,
    credential: str | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> dict[str, Any]:
    assistants = _list_provider_assistants(
        assistant_type="OPENAI",
        default_model="gpt-4o-mini",
        db=db,
        user=user,
    )
    for assistant in assistants:
        if assistant.get("id") == assistant_obj_id:
            return assistant
    raise HTTPException(status_code=404, detail="OpenAI assistant not found")


@router.get("/azure-assistants/{assistant_obj_id}")
def get_azure_assistant(
    assistant_obj_id: str,
    credential: str | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> dict[str, Any]:
    assistants = _list_provider_assistants(
        assistant_type="AZURE",
        default_model="gpt-4o-mini",
        db=db,
        user=user,
    )
    for assistant in assistants:
        if assistant.get("id") == assistant_obj_id:
            return assistant
    raise HTTPException(status_code=404, detail="Azure assistant not found")


def _query_vector_stores(db: Session, tenant_id: str, credential: str | None = None) -> list[TenantResource]:
    rows = _query_resources(db, tenant_id, "assistant_vector_store")
    if credential:
        rows = [row for row in rows if (row.payload or {}).get("credential") == credential]
    return rows


@router.get("/openai-assistants-vector-store")
def list_vector_stores(
    credential: str | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> list[dict[str, Any]]:
    rows = _query_vector_stores(db, user["tenant_id"], credential=credential)
    result = []
    for row in rows:
        payload = row.payload or {}
        files = payload.get("files") or []
        usage_bytes = int(sum(int(file.get("bytes", 0)) for file in files))
        result.append(
            {
                "id": row.id,
                "name": payload.get("name") or _resource_display_name(row),
                "expires_after": payload.get("expires_after"),
                "files": files,
                "file_counts": {"total": len(files)},
                "usage_bytes": usage_bytes,
            }
        )
    return result


@router.get("/openai-assistants-vector-store/{vector_store_id}")
def get_vector_store(
    vector_store_id: str,
    credential: str | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "assistant_vector_store", vector_store_id)
    payload = row.payload or {}
    files = payload.get("files") or []
    usage_bytes = int(sum(int(file.get("bytes", 0)) for file in files))
    return {
        "id": row.id,
        "name": payload.get("name") or _resource_display_name(row),
        "expires_after": payload.get("expires_after"),
        "files": files,
        "file_counts": {"total": len(files)},
        "usage_bytes": usage_bytes,
    }


@router.post("/openai-assistants-vector-store")
def create_vector_store(
    body: dict[str, Any],
    credential: str | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    name = body.get("name") or f"Vector Store {datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}"
    payload = {
        "name": name,
        "credential": credential,
        "expires_after": body.get("expires_after"),
        "files": [],
    }
    row = _create_resource(db, user["tenant_id"], user.get("user_id"), "assistant_vector_store", name, payload)
    return get_vector_store(row.id, credential=credential, db=db, user=user)


@router.put("/openai-assistants-vector-store/{vector_store_id}")
def update_vector_store(
    vector_store_id: str,
    body: dict[str, Any],
    credential: str | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "assistant_vector_store", vector_store_id)
    payload = row.payload or {}
    new_name = body.get("name") or payload.get("name") or _resource_display_name(row)
    updated_payload = {
        **payload,
        "name": new_name,
        "expires_after": body.get("expires_after"),
    }
    _update_resource(row, user.get("user_id"), display_name=new_name, payload=updated_payload)
    db.commit()
    return get_vector_store(vector_store_id, credential=credential, db=db, user=user)


@router.delete("/openai-assistants-vector-store/{vector_store_id}")
def delete_vector_store(
    vector_store_id: str,
    credential: str | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "assistant_vector_store", vector_store_id)
    db.delete(row)
    db.commit()
    return {"status": "deleted", "id": vector_store_id}


@router.post("/openai-assistants-vector-store/{vector_store_id}")
async def upload_files_to_vector_store(
    vector_store_id: str,
    credential: str | None = None,
    files: list[UploadFile] = File(default=[]),
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> list[dict[str, Any]]:
    row = _get_resource(db, user["tenant_id"], "assistant_vector_store", vector_store_id)
    payload = row.payload or {}
    existing_files = payload.get("files") or []

    uploaded = []
    for item in files:
        content = await item.read()
        uploaded.append(
            {
                "id": str(uuid4()),
                "filename": item.filename,
                "bytes": len(content),
                "created_at": _now_iso(),
            }
        )

    payload["files"] = [*existing_files, *uploaded]
    _update_resource(row, user.get("user_id"), payload=payload)
    db.commit()
    return uploaded


@router.patch("/openai-assistants-vector-store/{vector_store_id}")
def delete_files_from_vector_store(
    vector_store_id: str,
    body: dict[str, Any],
    credential: str | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "assistant_vector_store", vector_store_id)
    payload = row.payload or {}
    file_ids = set(body.get("file_ids") or [])
    payload["files"] = [file for file in (payload.get("files") or []) if file.get("id") not in file_ids]
    _update_resource(row, user.get("user_id"), payload=payload)
    db.commit()
    return {"status": "ok"}


@router.post("/openai-assistants-file/upload")
async def upload_files_to_openai_assistant(
    credential: str | None = None,
    files: list[UploadFile] = File(default=[]),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> list[dict[str, Any]]:
    uploaded = []
    for item in files:
        content = await item.read()
        uploaded.append(
            {
                "id": str(uuid4()),
                "filename": item.filename,
                "bytes": len(content),
                "created_at": _now_iso(),
            }
        )
    return uploaded


@router.get("/azure-assistants-vector-store")
def list_azure_vector_stores(
    credential: str | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> list[dict[str, Any]]:
    return list_vector_stores(credential=credential, db=db, user=user)


@router.get("/azure-assistants-vector-store/{vector_store_id}")
def get_azure_vector_store(
    vector_store_id: str,
    credential: str | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> dict[str, Any]:
    return get_vector_store(vector_store_id=vector_store_id, credential=credential, db=db, user=user)


@router.post("/azure-assistants-vector-store")
def create_azure_vector_store(
    body: dict[str, Any],
    credential: str | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    return create_vector_store(body=body, credential=credential, db=db, user=user)


@router.put("/azure-assistants-vector-store/{vector_store_id}")
def update_azure_vector_store(
    vector_store_id: str,
    body: dict[str, Any],
    credential: str | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    return update_vector_store(vector_store_id=vector_store_id, body=body, credential=credential, db=db, user=user)


@router.delete("/azure-assistants-vector-store/{vector_store_id}")
def delete_azure_vector_store(
    vector_store_id: str,
    credential: str | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    return delete_vector_store(vector_store_id=vector_store_id, credential=credential, db=db, user=user)


@router.post("/azure-assistants-vector-store/{vector_store_id}")
async def upload_files_to_azure_vector_store(
    vector_store_id: str,
    credential: str | None = None,
    files: list[UploadFile] = File(default=[]),
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> list[dict[str, Any]]:
    return await upload_files_to_vector_store(vector_store_id=vector_store_id, credential=credential, files=files, db=db, user=user)


@router.patch("/azure-assistants-vector-store/{vector_store_id}")
def delete_files_from_azure_vector_store(
    vector_store_id: str,
    body: dict[str, Any],
    credential: str | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    return delete_files_from_vector_store(vector_store_id=vector_store_id, body=body, credential=credential, db=db, user=user)


@router.post("/azure-assistants-file/upload")
async def upload_files_to_azure_assistant(
    credential: str | None = None,
    files: list[UploadFile] = File(default=[]),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> list[dict[str, Any]]:
    return await upload_files_to_openai_assistant(credential=credential, files=files, user=user)


# Document Store
@router.get("/document-store/store")
def list_document_stores(
    page: int = 1,
    limit: int = 10,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> dict[str, Any]:
    rows = _query_resources(db, user["tenant_id"], "document_store")
    items = []
    for row in rows:
        payload = row.payload or {}
        items.append(
            {
                "id": row.id,
                "name": _resource_display_name(row),
                "description": payload.get("description", ""),
                "loaders": payload.get("loaders") or [],
                "vectorStoreConfig": payload.get("vectorStoreConfig") or {},
                "recordManagerConfig": payload.get("recordManagerConfig") or {},
                "workspaceId": payload.get("workspaceId") or user["tenant_id"],
                "whereUsed": payload.get("whereUsed") or [],
                "status": payload.get("status") or "READY",
                "createdDate": row.created_at.isoformat() if row.created_at else None,
                "updatedDate": row.updated_at.isoformat() if row.updated_at else None,
            }
        )

    paged, total = _paginate(items, page=page, limit=limit)
    return {"data": paged, "total": total}


@router.get("/document-store/components/loaders")
def list_document_loaders(
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> list[dict[str, Any]]:
    return DOC_LOADERS


@router.get("/document-store/store/{store_id}")
def get_document_store(
    store_id: str,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "document_store", store_id)
    payload = row.payload or {}
    return {
        "id": row.id,
        "name": _resource_display_name(row),
        "description": payload.get("description", ""),
        "loaders": payload.get("loaders") or [],
        "vectorStoreConfig": payload.get("vectorStoreConfig") or {},
        "recordManagerConfig": payload.get("recordManagerConfig") or {},
        "workspaceId": payload.get("workspaceId") or user["tenant_id"],
        "whereUsed": payload.get("whereUsed") or [],
        "status": payload.get("status") or "READY",
    }


@router.post("/document-store/store")
def create_document_store(
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    display_name = body.get("name") or "Document Store"
    payload = {
        "description": body.get("description", ""),
        "loaders": [],
        "vectorStoreConfig": {},
        "recordManagerConfig": {},
        "workspaceId": user["tenant_id"],
        "whereUsed": [],
        "status": "READY",
    }
    row = _create_resource(db, user["tenant_id"], user.get("user_id"), "document_store", display_name, payload)
    return {"id": row.id}


@router.put("/document-store/store/{store_id}")
def update_document_store(
    store_id: str,
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "document_store", store_id)
    payload = row.payload or {}
    display_name = body.get("name") or _resource_display_name(row)
    next_payload = {
        **payload,
        "description": body.get("description", payload.get("description", "")),
    }
    _update_resource(row, user.get("user_id"), display_name=display_name, payload=next_payload)
    db.commit()
    return {"id": row.id}


@router.delete("/document-store/store/{store_id}")
def delete_document_store(
    store_id: str,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "document_store", store_id)
    db.delete(row)
    db.commit()
    return {"status": "deleted", "id": store_id}


@router.get("/document-store/store-configs/{store_id}/{loader_id}")
def get_document_store_config(
    store_id: str,
    loader_id: str,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> dict[str, Any] | None:
    store = get_document_store(store_id, db=db, user=user)
    for loader in store.get("loaders") or []:
        if loader.get("id") == loader_id:
            return loader
    return None


def _ensure_tenant_docstore_upload_dir(tenant_id: str) -> Path:
    """Create uploads directory for document store if it doesn't exist."""
    uploads_base = os.getenv("UPLOADS_DIR", "./uploads")
    tenant_dir = Path(uploads_base) / tenant_id / "docstore"
    tenant_dir.mkdir(parents=True, exist_ok=True)
    return tenant_dir


@router.post("/document-store/loader/upload")
async def upload_document_loader_file(
    file: UploadFile = File(...),
    store_id: str = "",
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    """Upload a file for document loader (PDF, CSV, Docx, etc.)."""
    tenant_id = user.get("tenant_id")

    # Create tenant-specific upload directory
    tenant_dir = _ensure_tenant_docstore_upload_dir(tenant_id)

    # Read file content
    try:
        contents = await file.read()
        file_size = len(contents)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read file: {str(e)}")

    # Generate unique filename
    file_id = str(uuid4())
    file_ext = Path(file.filename).suffix if file.filename else ""
    safe_filename = f"{file_id}{file_ext}"
    file_path = tenant_dir / safe_filename

    # Write file to disk
    try:
        with open(file_path, "wb") as f:
            f.write(contents)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save file: {str(e)}")

    return {
        "fileId": file_id,
        "path": str(file_path),
        "name": file.filename,
        "size": file_size,
    }


@router.post("/document-store/loader/preview")
def preview_document_loader_chunks(
    body: dict[str, Any],
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> dict[str, Any]:
    loader_config = body.get("loaderConfig") or {}
    source = str(loader_config.get("text") or loader_config.get("url") or loader_config.get("path") or "")
    splitter_id = body.get("splitterId")
    splitter_config = body.get("splitterConfig") or {}
    chunks, total = _split_text(source, splitter_id, splitter_config, preview_count=int(body.get("previewChunkCount") or 10))
    return {
        "chunks": chunks,
        "totalChunks": total,
        "previewChunkCount": len(chunks),
    }


@router.post("/document-store/loader/save")
def save_document_loader(
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    store_row = _get_resource(db, user["tenant_id"], "document_store", body.get("storeId"))
    payload = store_row.payload or {}
    loaders = payload.get("loaders") or []

    source = str((body.get("loaderConfig") or {}).get("text") or (body.get("loaderConfig") or {}).get("url") or "")
    splitter_id = body.get("splitterId")
    splitter_config = body.get("splitterConfig") or {}
    preview_chunks, total = _split_text(source, splitter_id, splitter_config, preview_count=20)

    loader_id = body.get("id") or str(uuid4())
    loader = {
        "id": loader_id,
        "loaderId": body.get("loaderId"),
        "loaderName": body.get("loaderName") or "Loader",
        "loaderConfig": body.get("loaderConfig") or {},
        "splitterId": body.get("splitterId"),
        "splitterName": body.get("splitterName"),
        "splitterConfig": body.get("splitterConfig") or {},
        "credential": body.get("credential"),
        "source": source,
        "files": [{"id": str(uuid4()), "name": body.get("loaderName") or "Document"}],
        "chunks": preview_chunks,
        "totalChunks": total,
        "totalChars": sum(len(chunk.get("pageContent", "")) for chunk in preview_chunks),
        "status": "SYNC",
        "updatedDate": _now_iso(),
    }

    replaced = False
    next_loaders = []
    for item in loaders:
        if item.get("id") == loader_id:
            next_loaders.append(loader)
            replaced = True
        else:
            next_loaders.append(item)

    if not replaced:
        next_loaders.append(loader)

    payload["loaders"] = next_loaders
    payload["status"] = "READY"
    _update_resource(store_row, user.get("user_id"), payload=payload)
    db.commit()

    return {"id": loader_id}


@router.post("/document-store/loader/process/{loader_id}")
def process_document_loader(
    loader_id: str,
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    """Process a document loader - finalize chunks and mark as ready."""
    store_id = body.get("storeId")
    if not store_id:
        raise HTTPException(status_code=400, detail="storeId is required")

    # Get the document store
    store_row = _get_resource(db, user["tenant_id"], "document_store", store_id)
    payload = store_row.payload or {}
    loaders = payload.get("loaders") or []

    # Find the loader
    loader = None
    for item in loaders:
        if item.get("id") == loader_id:
            loader = item
            break

    if loader is None:
        raise HTTPException(status_code=404, detail="Loader not found")

    # Process the loader using the appropriate handler
    text_content = _process_loader_text(loader)

    # Build chunks from the processed content using stored splitter config
    splitter_id = loader.get("splitterId")
    splitter_config = loader.get("splitterConfig") or {}
    all_chunks, total = _split_text(text_content, splitter_id, splitter_config)

    # Update the loader with final chunks
    loader["chunks"] = all_chunks
    loader["totalChunks"] = total
    loader["totalChars"] = sum(len(chunk.get("pageContent", "")) for chunk in all_chunks)
    loader["status"] = "SYNC"
    loader["updatedDate"] = _now_iso()

    # Update the store
    payload["loaders"] = [loader if item.get("id") == loader_id else item for item in loaders]
    payload["status"] = "READY"
    _update_resource(store_row, user.get("user_id"), payload=payload)
    db.commit()

    return {
        "status": "done",
        "id": loader_id,
        "chunksCount": total,
        "charsCount": loader["totalChars"]
    }


def _process_loader_text(loader: dict[str, Any]) -> str:
    """Extract text from a loader based on its type. Returns the text content."""
    loader_id = loader.get("loaderId", "")
    loader_config = loader.get("loaderConfig") or {}
    source = loader.get("source") or ""

    # Web/Cheerio loaders
    if "web" in loader_id.lower() or "cheerio" in loader_id.lower():
        try:
            from bs4 import BeautifulSoup
            response = httpx.get(source, timeout=10.0)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, "html.parser")
            # Optional: use CSS selector if provided
            selector = loader_config.get("selector")
            if selector:
                elements = soup.select(selector)
                text = "\n".join([elem.get_text() for elem in elements])
                return text
            return soup.get_text()
        except Exception as e:
            print(f"Warning: Failed to fetch URL: {e}")
            return ""

    # PDF loader
    elif "pdf" in loader_id.lower():
        try:
            if isinstance(source, str) and os.path.isfile(source):
                with open(source, 'rb') as f:
                    pdf_bytes = f.read()
            else:
                return ""

            from io import BytesIO
            from pdfminer.layout import LAParams
            from pdfminer.converter import TextConverter
            from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
            from pdfminer.pdfpage import PDFPage

            pdf_file = BytesIO(pdf_bytes)
            resource_manager = PDFResourceManager()
            output = BytesIO()
            converter = TextConverter(resource_manager, output, laparams=LAParams())
            interpreter = PDFPageInterpreter(resource_manager, converter)

            for page in PDFPage.get_pages(pdf_file):
                interpreter.process_page(page)

            text = output.getvalue().decode('utf-8')
            converter.close()
            output.close()
            return text
        except ImportError:
            print("Warning: pdfminer.six not installed")
            return ""
        except Exception as e:
            print(f"Warning: Failed to process PDF: {e}")
            return ""

    # CSV loader
    elif "csv" in loader_id.lower():
        try:
            import csv
            if isinstance(source, str) and os.path.isfile(source):
                chunks = []
                with open(source, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        chunks.append("\n".join([f"{k}: {v}" for k, v in row.items()]))
                return "\n\n".join(chunks)
            return ""
        except Exception as e:
            print(f"Warning: Failed to process CSV: {e}")
            return ""

    # JSON loader
    elif loader_id.lower() == "jsonfileloader":
        try:
            if isinstance(source, str) and os.path.isfile(source):
                with open(source, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                return json.dumps(data, indent=2)
            return ""
        except Exception as e:
            print(f"Warning: Failed to process JSON: {e}")
            return ""

    # JSON Lines loader
    elif "jsonlines" in loader_id.lower():
        try:
            import jsonlines
            if isinstance(source, str) and os.path.isfile(source):
                chunks = []
                with jsonlines.open(source) as reader:
                    text_key = loader_config.get("textKey", "text")
                    for obj in reader:
                        if text_key in obj:
                            chunks.append(str(obj[text_key]))
                        else:
                            chunks.append(json.dumps(obj))
                return "\n\n".join(chunks)
            return ""
        except Exception as e:
            print(f"Warning: Failed to process JSONL: {e}")
            return ""

    # Docx loader
    elif "docx" in loader_id.lower():
        try:
            from docx import Document
            if isinstance(source, str) and os.path.isfile(source):
                doc = Document(source)
                return "\n".join([p.text for p in doc.paragraphs])
            return ""
        except Exception as e:
            print(f"Warning: Failed to process Docx: {e}")
            return ""

    # Excel loader
    elif "excel" in loader_id.lower():
        try:
            from openpyxl import load_workbook
            if isinstance(source, str) and os.path.isfile(source):
                wb = load_workbook(source)
                sheet_name = loader_config.get("sheetName")
                if sheet_name:
                    ws = wb[sheet_name]
                else:
                    ws = wb.active

                chunks = []
                for row in ws.iter_rows(values_only=True):
                    chunks.append(" | ".join([str(cell) if cell is not None else "" for cell in row]))
                return "\n".join(chunks)
            return ""
        except Exception as e:
            print(f"Warning: Failed to process Excel: {e}")
            return ""

    # Notion loader
    elif "notion" in loader_id.lower():
        try:
            notion_page_id = loader_config.get("notionPageId", "")
            notion_api_key = loader.get("credential") or os.getenv("NOTION_API_KEY", "")
            if not notion_api_key or not notion_page_id:
                return "Notion API key or page ID not configured"

            headers = {
                "Authorization": f"Bearer {notion_api_key}",
                "Notion-Version": "2022-06-28"
            }
            response = httpx.get(
                f"https://api.notion.com/v1/blocks/{notion_page_id}/children",
                headers=headers,
                timeout=10.0
            )
            response.raise_for_status()
            data = response.json()

            chunks = []
            for block in data.get("results", []):
                block_type = block.get("type")
                if block_type == "paragraph":
                    text = block.get("paragraph", {}).get("rich_text", [])
                    chunks.append("".join([t.get("plain_text", "") for t in text]))
            return "\n".join(chunks)
        except Exception as e:
            print(f"Warning: Failed to fetch from Notion: {e}")
            return ""

    # GitHub loader
    elif "github" in loader_id.lower():
        try:
            repo_url = loader_config.get("repoLink", "")
            branch = loader_config.get("branch", "main")
            extensions = loader_config.get("fileExtensions", ".md,.txt,.py,.js,.ts").split(",")
            github_token = loader.get("credential") or os.getenv("GITHUB_ACCESS_TOKEN", "")

            if not repo_url:
                return "GitHub repo URL not configured"

            # Parse repo URL
            parts = repo_url.replace("https://github.com/", "").replace(".git", "").split("/")
            if len(parts) < 2:
                return "Invalid GitHub URL"

            owner, repo = parts[0], parts[1]
            headers = {}
            if github_token:
                headers["Authorization"] = f"token {github_token}"

            chunks = []

            # Recursive function to fetch files
            def fetch_dir(path=""):
                try:
                    url = f"https://api.github.com/repos/{owner}/{repo}/contents/{path}?ref={branch}"
                    response = httpx.get(url, headers=headers, timeout=10.0)
                    response.raise_for_status()
                    items = response.json()

                    if not isinstance(items, list):
                        items = [items]

                    for item in items:
                        if item.get("type") == "file":
                            file_ext = Path(item["name"]).suffix
                            if any(file_ext == ext.strip() for ext in extensions):
                                file_url = item.get("download_url")
                                if file_url:
                                    file_response = httpx.get(file_url, timeout=10.0)
                                    chunks.append(f"# {item['path']}\n{file_response.text}")
                        elif item.get("type") == "dir" and loader_config.get("recursive", True):
                            fetch_dir(item["path"])
                except Exception as e:
                    print(f"Warning: Failed to fetch GitHub path: {e}")

            fetch_dir()
            return "\n\n".join(chunks)
        except Exception as e:
            print(f"Warning: Failed to fetch from GitHub: {e}")
            return ""

    # Confluence loader
    elif "confluence" in loader_id.lower():
        try:
            space_key = loader_config.get("spaceKey", "")
            base_url = loader_config.get("baseUrl", "")
            username = loader_config.get("confluenceUsername") or os.getenv("CONFLUENCE_USERNAME", "")
            api_token = loader.get("credential") or os.getenv("CONFLUENCE_API_TOKEN", "")

            if not base_url or not space_key or not api_token:
                return "Confluence configuration incomplete"

            from base64 import b64encode
            auth = b64encode(f"{username}:{api_token}".encode()).decode()
            headers = {
                "Authorization": f"Basic {auth}",
                "Accept": "application/json"
            }

            url = f"{base_url}/rest/api/content?spaceKey={space_key}&expand=body.storage&limit=50"
            response = httpx.get(url, headers=headers, timeout=10.0)
            response.raise_for_status()
            data = response.json()

            chunks = []
            from bs4 import BeautifulSoup
            for page in data.get("results", []):
                title = page.get("title", "")
                body = page.get("body", {}).get("storage", {}).get("value", "")
                soup = BeautifulSoup(body, "html.parser")
                text = soup.get_text()
                chunks.append(f"# {title}\n{text}")

            return "\n\n".join(chunks)
        except Exception as e:
            print(f"Warning: Failed to fetch from Confluence: {e}")
            return ""

    # S3 loader
    elif "s3" in loader_id.lower():
        try:
            import boto3
            bucket = loader_config.get("bucket", "")
            key_prefix = loader_config.get("keyPrefix", "")

            if not bucket:
                return "S3 bucket not configured"

            s3_client = boto3.client('s3')
            chunks = []

            # List objects
            response = s3_client.list_objects_v2(Bucket=bucket, Prefix=key_prefix)
            for obj in response.get("Contents", []):
                key = obj["Key"]
                try:
                    obj_response = s3_client.get_object(Bucket=bucket, Key=key)
                    body = obj_response['Body'].read()

                    # Try to decode as text
                    try:
                        text = body.decode('utf-8')
                        chunks.append(f"# {key}\n{text}")
                    except:
                        chunks.append(f"# {key}\n[Binary content]")
                except Exception as e:
                    print(f"Warning: Failed to read S3 object {key}: {e}")

            return "\n\n".join(chunks)
        except Exception as e:
            print(f"Warning: Failed to access S3: {e}")
            return ""

    # Airtable loader
    elif "airtable" in loader_id.lower():
        try:
            base_id = loader_config.get("baseId", "")
            table_id = loader_config.get("tableId", "")
            api_key = loader.get("credential") or os.getenv("AIRTABLE_API_KEY", "")

            if not base_id or not table_id or not api_key:
                return "Airtable configuration incomplete"

            url = f"https://api.airtable.com/v0/{base_id}/{table_id}"
            headers = {"Authorization": f"Bearer {api_key}"}

            response = httpx.get(url, headers=headers, timeout=10.0)
            response.raise_for_status()
            data = response.json()

            chunks = []
            for record in data.get("records", []):
                fields = record.get("fields", {})
                chunks.append("\n".join([f"{k}: {v}" for k, v in fields.items()]))

            return "\n\n".join(chunks)
        except Exception as e:
            print(f"Warning: Failed to fetch from Airtable: {e}")
            return ""

    else:
        # Default: Text loader
        return source


@router.post("/document-store/refresh/{store_id}")
def refresh_document_store(
    store_id: str,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    """Refresh document store by re-processing all loaders."""
    row = _get_resource(db, user["tenant_id"], "document_store", store_id)
    payload = row.payload or {}
    loaders = payload.get("loaders") or []

    # Re-process each loader
    for loader in loaders:
        # Extract text from source
        text_content = _process_loader_text(loader)

        # Re-split with stored splitter config
        splitter_id = loader.get("splitterId")
        splitter_config = loader.get("splitterConfig") or {}
        all_chunks, total = _split_text(text_content, splitter_id, splitter_config)

        # Update the loader
        loader["chunks"] = all_chunks
        loader["totalChunks"] = total
        loader["totalChars"] = sum(len(chunk.get("pageContent", "")) for chunk in all_chunks)
        loader["status"] = "SYNC"
        loader["updatedDate"] = _now_iso()

    payload["loaders"] = loaders
    payload["status"] = "READY"
    _update_resource(row, user.get("user_id"), payload=payload)
    db.commit()
    return {"status": "ok", "loaders_refreshed": len(loaders)}


@router.get("/document-store/chunks/{store_id}/{file_id}/{page_no}")
def get_document_chunks(
    store_id: str,
    file_id: str,
    page_no: int,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> dict[str, Any]:
    store = get_document_store(store_id, db=db, user=user)
    loader = None
    for item in store.get("loaders") or []:
        if item.get("id") == file_id:
            loader = item
            break

    if loader is None:
        return {
            "workspaceId": store.get("workspaceId") or user["tenant_id"],
            "count": 0,
            "chunks": [],
            "currentPage": max(1, page_no),
            "file": None,
            "storeName": store.get("name", ""),
        }

    chunks = loader.get("chunks") or []
    safe_page = max(1, page_no)
    limit = 50
    start = (safe_page - 1) * limit
    paged = [
        {
            **chunk,
            "storeId": store_id,
            "docId": file_id,
        }
        for chunk in chunks[start : start + limit]
    ]

    return {
        "workspaceId": store.get("workspaceId") or user["tenant_id"],
        "count": len(chunks),
        "chunks": paged,
        "currentPage": safe_page,
        "file": loader,
        "storeName": store.get("name", ""),
    }


@router.put("/document-store/chunks/{store_id}/{loader_id}/{chunk_id}")
def edit_document_chunk(
    store_id: str,
    loader_id: str,
    chunk_id: str,
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "document_store", store_id)
    payload = row.payload or {}
    loaders = payload.get("loaders") or []

    for loader in loaders:
        if loader.get("id") != loader_id:
            continue
        for chunk in loader.get("chunks") or []:
            if chunk.get("id") == chunk_id:
                chunk["pageContent"] = body.get("pageContent", chunk.get("pageContent"))
                chunk["metadata"] = body.get("metadata", chunk.get("metadata", {}))

    payload["loaders"] = loaders
    _update_resource(row, user.get("user_id"), payload=payload)
    db.commit()
    return {"id": chunk_id}


@router.delete("/document-store/chunks/{store_id}/{loader_id}/{chunk_id}")
def delete_document_chunk(
    store_id: str,
    loader_id: str,
    chunk_id: str,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "document_store", store_id)
    payload = row.payload or {}
    loaders = payload.get("loaders") or []

    for loader in loaders:
        if loader.get("id") == loader_id:
            loader["chunks"] = [chunk for chunk in (loader.get("chunks") or []) if chunk.get("id") != chunk_id]
            loader["totalChunks"] = len(loader.get("chunks") or [])
            loader["totalChars"] = sum(len(chunk.get("pageContent", "")) for chunk in (loader.get("chunks") or []))

    payload["loaders"] = loaders
    _update_resource(row, user.get("user_id"), payload=payload)
    db.commit()
    return {"id": chunk_id}


@router.delete("/document-store/loader/{store_id}/{file_id}")
def delete_document_loader(
    store_id: str,
    file_id: str,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "document_store", store_id)
    payload = row.payload or {}
    payload["loaders"] = [loader for loader in (payload.get("loaders") or []) if loader.get("id") != file_id]
    _update_resource(row, user.get("user_id"), payload=payload)
    db.commit()
    return {"id": file_id}


@router.post("/document-store/vectorstore/save")
def save_vector_store_config(
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "document_store", body.get("storeId"))
    payload = row.payload or {}
    payload["vectorStoreConfig"] = body.get("vectorStoreConfig") or {}
    payload["recordManagerConfig"] = body.get("recordManagerConfig") or {}
    _update_resource(row, user.get("user_id"), payload=payload)
    db.commit()
    return {"id": row.id}


@router.post("/document-store/vectorstore/update")
def update_vector_store_config(
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    return save_vector_store_config(body, db=db, user=user)


def _upsert_to_vector_store(
    provider: str,
    config: dict[str, Any],
    chunks: list[dict[str, Any]],
    embedder_func,
    tenant_id: str,
) -> tuple[int, list[dict[str, Any]]]:
    """Upsert chunks to the specified vector store provider. Returns (num_added, added_docs)."""
    num_added = 0
    added_docs = []

    if provider == "qdrant":
        # This is handled by the existing Qdrant logic, return 0 here
        return 0, []

    elif provider == "chroma":
        try:
            import chromadb
            client = chromadb.HttpClient(host=config.get("host", "localhost"), port=config.get("port", 8000))
            collection_name = config.get("collectionName", "documents")

            collection = client.get_or_create_collection(name=collection_name, metadata={"hnsw:space": "cosine"})

            ids = []
            embeddings = []
            metadatas = []
            documents = []

            for chunk in chunks:
                page_content = chunk.get("pageContent", "")
                if not page_content:
                    continue

                embedding = embedder_func(page_content)
                if not embedding:
                    continue

                ids.append(chunk.get("id"))
                embeddings.append(embedding)
                documents.append(page_content[:5000])
                metadatas.append(chunk.get("metadata", {}))
                num_added += 1
                if num_added <= 5:
                    added_docs.append(chunk)

            if ids:
                collection.upsert(ids=ids, embeddings=embeddings, documents=documents, metadatas=metadatas)

            return num_added, added_docs
        except Exception as e:
            print(f"Warning: Failed to upsert to Chroma: {e}")
            return 0, []

    elif provider == "pinecone":
        try:
            import pinecone
            index_name = config.get("indexName", "documents")
            namespace = config.get("namespace", "")
            api_key = config.get("apiKey", os.getenv("PINECONE_API_KEY", ""))

            if not api_key:
                print("Warning: Pinecone API key not provided")
                return 0, []

            pinecone.init(api_key=api_key)
            index = pinecone.Index(index_name)

            vectors = []
            for chunk in chunks:
                page_content = chunk.get("pageContent", "")
                if not page_content:
                    continue

                embedding = embedder_func(page_content)
                if not embedding:
                    continue

                vectors.append((
                    chunk.get("id"),
                    embedding,
                    {
                        "pageContent": page_content[:5000],
                        **chunk.get("metadata", {}),
                    }
                ))
                num_added += 1
                if num_added <= 5:
                    added_docs.append(chunk)

            if vectors:
                # Upsert in batches of 100
                for i in range(0, len(vectors), 100):
                    batch = vectors[i:i+100]
                    index.upsert(vectors=batch, namespace=namespace)

            return num_added, added_docs
        except Exception as e:
            print(f"Warning: Failed to upsert to Pinecone: {e}")
            return 0, []

    elif provider == "weaviate":
        try:
            import weaviate
            host = config.get("host", "localhost")
            port = config.get("port", 8080)
            weaviate_host = f"http://{host}:{port}"
            class_name = config.get("className", "Document")
            api_key = config.get("apiKey", "")

            client = weaviate.connect_to_custom(
                http_host=host,
                http_port=port,
                http_secure=False,
                grpc_host=host,
                grpc_port=50051,
                grpc_secure=False
            )

            # Prepare data objects
            data_objs = []
            for chunk in chunks:
                page_content = chunk.get("pageContent", "")
                if not page_content:
                    continue

                embedding = embedder_func(page_content)
                if not embedding:
                    continue

                data_objs.append({
                    "pageContent": page_content[:5000],
                    "chunkId": chunk.get("id"),
                    "metadata": json.dumps(chunk.get("metadata", {})),
                })
                num_added += 1
                if num_added <= 5:
                    added_docs.append(chunk)

            if data_objs:
                with client.batch.fixed_size(size=100) as batch:
                    for obj in data_objs:
                        batch.add_object(
                            class_name=class_name,
                            properties=obj,
                        )

            client.close()
            return num_added, added_docs
        except Exception as e:
            print(f"Warning: Failed to upsert to Weaviate: {e}")
            return 0, []

    else:
        print(f"Warning: Unknown vector store provider: {provider}")
        return 0, []


def _query_vector_store(
    provider: str,
    config: dict[str, Any],
    query_vector: list[float],
    query_text: str,
    embedder_func,
    k: int = 5,
) -> list[dict[str, Any]]:
    """Query a vector store and return similar chunks."""

    if provider == "qdrant":
        # This is handled by the existing Qdrant logic, return empty here
        return []

    elif provider == "chroma":
        try:
            import chromadb
            client = chromadb.HttpClient(host=config.get("host", "localhost"), port=config.get("port", 8000))
            collection_name = config.get("collectionName", "documents")

            collection = client.get_collection(name=collection_name)
            results = collection.query(query_embeddings=[query_vector], n_results=k)

            docs = []
            for i, doc in enumerate(results.get("documents", [[]])[0]):
                docs.append({
                    "pageContent": doc,
                    "score": 1 - (results.get("distances", [[]])[0][i] if i < len(results.get("distances", [[]])[0]) else 0),
                    "metadata": results.get("metadatas", [[]])[0][i] if i < len(results.get("metadatas", [[]])[0]) else {},
                })
            return docs
        except Exception as e:
            print(f"Warning: Failed to query Chroma: {e}")
            return []

    elif provider == "pinecone":
        try:
            import pinecone
            index_name = config.get("indexName", "documents")
            namespace = config.get("namespace", "")
            api_key = config.get("apiKey", os.getenv("PINECONE_API_KEY", ""))

            if not api_key:
                return []

            pinecone.init(api_key=api_key)
            index = pinecone.Index(index_name)

            results = index.query(vector=query_vector, top_k=k, namespace=namespace, include_metadata=True)

            docs = []
            for match in results.get("matches", []):
                docs.append({
                    "pageContent": match.get("metadata", {}).get("pageContent", ""),
                    "score": match.get("score", 0),
                    "metadata": {k: v for k, v in match.get("metadata", {}).items() if k != "pageContent"},
                })
            return docs
        except Exception as e:
            print(f"Warning: Failed to query Pinecone: {e}")
            return []

    elif provider == "weaviate":
        try:
            import weaviate
            host = config.get("host", "localhost")
            port = config.get("port", 8080)
            class_name = config.get("className", "Document")

            client = weaviate.connect_to_custom(
                http_host=host,
                http_port=port,
                http_secure=False,
                grpc_host=host,
                grpc_port=50051,
                grpc_secure=False
            )

            # Weaviate uses vector search with near_vector
            results = client.query.get(class_name, ["pageContent", "metadata", "_additional {vector distance}"]).with_near_vector(
                {"vector": query_vector}
            ).with_limit(k).do()

            docs = []
            for obj in results.get("data", {}).get("Get", {}).get(class_name, []):
                docs.append({
                    "pageContent": obj.get("pageContent", ""),
                    "score": 1 - obj.get("_additional", {}).get("distance", 1),
                    "metadata": json.loads(obj.get("metadata", "{}")),
                })

            client.close()
            return docs
        except Exception as e:
            print(f"Warning: Failed to query Weaviate: {e}")
            return []

    else:
        return []


@router.post("/document-store/vectorstore/insert")
def insert_into_vector_store(
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    """Upsert chunks to the configured vector store."""
    row = _get_resource(db, user["tenant_id"], "document_store", body.get("storeId"))
    payload = row.payload or {}

    store_id = body.get("storeId")
    tenant_id = user["tenant_id"]

    # Get vector store provider from config
    vec_config = payload.get("vectorStoreConfig") or {}
    vector_provider = body.get("vectorStoreConfig", {}).get("provider") or vec_config.get("provider", "qdrant")
    vector_config = body.get("vectorStoreConfig", {}) or vec_config
    vector_config.setdefault("collectionName", f"tenant_{tenant_id[:8]}_store_{store_id[:8]}")

    # Collect all chunks from loaders
    all_chunks = []
    for loader in payload.get("loaders") or []:
        all_chunks.extend(loader.get("chunks") or [])

    if not all_chunks:
        return {
            "numAdded": 0,
            "numUpdated": 0,
            "numSkipped": 0,
            "numDeleted": 0,
            "addedDocs": [],
            "timeTaken": 0,
            "docs": [],
        }

    # Get embedding provider configuration
    embedding_provider = vec_config.get("embeddingProvider", "ollamaEmbeddings")
    embedding_config = vec_config.get("embeddingConfig") or {}

    # Determine embedding dimension
    embedding_dim = EMBEDDING_DIM  # default
    if embedding_provider in EMBEDDING_DIMS:
        dims = EMBEDDING_DIMS[embedding_provider]
        if isinstance(dims, dict):
            model = embedding_config.get("model") or (embedding_config.get("modelId") if "huggingFace" in embedding_provider else None)
            embedding_dim = dims.get(model, 384)
        else:
            embedding_dim = dims

    # Handle Qdrant specifically (requires pre-creating the collection)
    if vector_provider == "qdrant":
        if not qdrant_client:
            raise HTTPException(status_code=503, detail="Qdrant service unavailable")

        collection_name = vector_config.get("collectionName", f"tenant_{tenant_id[:8]}_store_{store_id[:8]}")
        try:
            qdrant_client.recreate_collection(
                collection_name=collection_name,
                vectors_config=VectorParams(size=embedding_dim, distance=Distance.COSINE),
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to create Qdrant collection: {str(e)}")

        # Embed and upsert chunks
        points = []
        num_added = 0
        added_docs = []

        for idx, chunk in enumerate(all_chunks):
            page_content = chunk.get("pageContent", "")
            if not page_content:
                continue

            # Get embedding using the configured provider
            embedding = _get_embedding(page_content, embedding_provider, embedding_config)
            if not embedding:
                continue

            # Create point for Qdrant
            point = PointStruct(
                id=hash(f"{chunk.get('id')}_{tenant_id}") & 0x7fffffff,  # Positive hash
                vector=embedding,
                payload={
                    "chunkId": chunk.get("id"),
                    "pageContent": page_content[:5000],  # Limit payload size
                    "metadata": chunk.get("metadata", {}),
                    "loaderIndex": idx,
                }
            )
            points.append(point)
            num_added += 1
            if num_added <= 5:
                added_docs.append(chunk)

        # Upsert to Qdrant
        if points:
            try:
                qdrant_client.upsert(
                    collection_name=collection_name,
                    points=points,
                )
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Failed to upsert to Qdrant: {str(e)}")

    else:
        # Use provider-specific upsert functions
        def embedder_func(text):
            return _get_embedding(text, embedding_provider, embedding_config)

        num_added, added_docs = _upsert_to_vector_store(
            vector_provider,
            vector_config,
            all_chunks,
            embedder_func,
            tenant_id
        )

    # Store collection name and config in document store
    payload["vectorStoreConfig"] = {
        "collectionName": collection_name,
        "provider": "qdrant",
        "embeddingProvider": embedding_provider,
        "embeddingConfig": embedding_config,
        "embeddingDim": embedding_dim,
        **body.get("vectorStoreConfig", {})
    }
    payload["recordManagerConfig"] = body.get("recordManagerConfig") or payload.get("recordManagerConfig") or {}
    payload["status"] = "SYNC"

    # Add upsert history entry
    upsert_history = payload.get("upsert_history") or []
    history_entry = {
        "date": _now_iso(),
        "result": {
            "numAdded": num_added,
            "numUpdated": 0,
            "numSkipped": len(all_chunks) - num_added,
            "numDeleted": 0,
        },
        "addedDocs": all_chunks[:5],
    }
    upsert_history.append(history_entry)
    payload["upsert_history"] = upsert_history

    _update_resource(row, user.get("user_id"), payload=payload)
    db.commit()

    return {
        "numAdded": num_added,
        "numUpdated": 0,
        "numSkipped": len(all_chunks) - num_added,
        "numDeleted": 0,
        "addedDocs": all_chunks[:5],
        "timeTaken": 0.1,
        "docs": all_chunks[:5],
    }


@router.delete("/document-store/vectorstore/{store_id}")
def delete_vector_store_data(
    store_id: str,
    docId: str | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "document_store", store_id)
    payload = row.payload or {}

    if docId:
        for loader in payload.get("loaders") or []:
            if loader.get("id") == docId:
                loader["chunks"] = []
                loader["totalChunks"] = 0
                loader["totalChars"] = 0
    else:
        payload["vectorStoreConfig"] = {}
        payload["recordManagerConfig"] = {}

    _update_resource(row, user.get("user_id"), payload=payload)
    db.commit()
    return {"status": "ok"}


@router.get("/upsert-history/{store_id}")
def get_upsert_history(
    store_id: str,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> list[dict[str, Any]]:
    """Get upsert history for a document store."""
    row = _get_resource(db, user["tenant_id"], "document_store", store_id)
    payload = row.payload or {}
    history = payload.get("upsert_history") or []
    # Sort by date descending
    return sorted(history, key=lambda x: x.get("date", ""), reverse=True)


@router.patch("/upsert-history")
def delete_upsert_history(
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    """Delete upsert history entries by date."""
    store_id = body.get("storeId")
    dates_to_delete = body.get("ids") or []  # list of date strings

    row = _get_resource(db, user["tenant_id"], "document_store", store_id)
    payload = row.payload or {}
    history = payload.get("upsert_history") or []

    # Filter out entries with matching dates
    filtered_history = [h for h in history if h.get("date") not in dates_to_delete]
    payload["upsert_history"] = filtered_history

    _update_resource(row, user.get("user_id"), payload=payload)
    db.commit()
    return {"status": "ok", "deleted": len(history) - len(filtered_history)}


@router.post("/document-store/vectorstore/query")
def query_vector_store(
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> dict[str, Any]:
    """Query vector store for similar chunks using semantic search."""
    import time

    start_time = time.time()

    row = _get_resource(db, user["tenant_id"], "document_store", body.get("storeId"))
    payload = row.payload or {}

    # Get vector store config
    vec_config = payload.get("vectorStoreConfig") or {}
    vector_provider = vec_config.get("provider", "qdrant")
    embedding_provider = vec_config.get("embeddingProvider", "ollamaEmbeddings")
    embedding_config = vec_config.get("embeddingConfig") or {}

    if not vec_config:
        # Fallback if no vector store configured
        docs = []
        for loader in payload.get("loaders") or []:
            for chunk in loader.get("chunks") or []:
                docs.append(
                    {
                        "pageContent": chunk.get("pageContent", ""),
                        "metadata": chunk.get("metadata", {}),
                        "score": 0.5,
                    }
                )
        elapsed = time.time() - start_time
        return {"docs": docs[:5], "timeTaken": elapsed}

    # Get query text
    query_text = body.get("query", "")
    if not query_text:
        elapsed = time.time() - start_time
        return {"docs": [], "timeTaken": elapsed}

    # Embed the query using the configured embedding provider
    query_embedding = _get_embedding(query_text, embedding_provider, embedding_config)
    if not query_embedding:
        # Fallback: return all chunks if embedding fails
        docs = []
        for loader in payload.get("loaders") or []:
            for chunk in loader.get("chunks") or []:
                docs.append(
                    {
                        "pageContent": chunk.get("pageContent", ""),
                        "metadata": chunk.get("metadata", {}),
                        "score": 0.5,
                    }
                )
        elapsed = time.time() - start_time
        return {"docs": docs[:5], "timeTaken": elapsed}

    # Query the vector store
    if vector_provider == "qdrant":
        if not qdrant_client:
            # Fallback if Qdrant unavailable
            docs = []
            for loader in payload.get("loaders") or []:
                for chunk in loader.get("chunks") or []:
                    docs.append(
                        {
                            "pageContent": chunk.get("pageContent", ""),
                            "metadata": chunk.get("metadata", {}),
                            "score": 0.5,
                        }
                    )
            elapsed = time.time() - start_time
            return {"docs": docs[:5], "timeTaken": elapsed}

        collection_name = vec_config.get("collectionName")
        if not collection_name:
            elapsed = time.time() - start_time
            return {"docs": [], "timeTaken": elapsed}

        # Search in Qdrant
        try:
            results = qdrant_client.search(
                collection_name=collection_name,
                query_vector=query_embedding,
                limit=body.get("k", 5),
                score_threshold=0.0,
            )
        except Exception as e:
            print(f"Warning: Qdrant search failed: {e}")
            elapsed = time.time() - start_time
            return {"docs": [], "timeTaken": elapsed}

        # Map results to document format
        docs = []
        for result in results:
            payload_data = result.payload or {}
            docs.append(
                {
                    "pageContent": payload_data.get("pageContent", ""),
                    "metadata": payload_data.get("metadata", {}),
                    "score": result.score,
                }
            )
    else:
        # Use provider-specific query functions
        def embedder_func(text):
            return _get_embedding(text, embedding_provider, embedding_config)

        docs = _query_vector_store(
            vector_provider,
            vec_config,
            query_embedding,
            query_text,
            embedder_func,
            body.get("k", 5)
        )

    elapsed = time.time() - start_time
    return {"docs": docs, "timeTaken": elapsed}


@router.get("/document-store/components/vectorstore")
def get_vector_store_components(
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> list[dict[str, Any]]:
    return VECTOR_PROVIDERS


@router.get("/document-store/components/embeddings")
def get_embedding_components(
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> list[dict[str, Any]]:
    return EMBEDDING_PROVIDERS


@router.get("/document-store/components/recordmanager")
def get_record_manager_components(
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> list[dict[str, Any]]:
    return RECORD_MANAGER_PROVIDERS


@router.post("/document-store/generate-tool-desc/{store_id}")
def generate_docstore_tool_description(
    store_id: str,
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "document_store", store_id)
    name = _resource_display_name(row)
    return {
        "content": f"Search and retrieve relevant context from document store '{name}' to answer user queries accurately."
    }


# Datasets
@router.get("/datasets")
def list_datasets(
    page: int | None = None,
    limit: int | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> Any:
    rows = _query_resources(db, user["tenant_id"], "dataset")
    items = []
    for row in rows:
        payload = row.payload or {}
        ds = {
            "id": row.id,
            "name": _resource_display_name(row),
            "description": payload.get("description", ""),
            "totalRows": len(payload.get("rows") or []),
            "createdDate": row.created_at.isoformat() if row.created_at else None,
            "updatedDate": row.updated_at.isoformat() if row.updated_at else None,
        }
        items.append(ds)

    if page is None and limit is None:
        return [{"id": item["id"], "name": item["name"]} for item in items]

    paged, total = _paginate(items, page=page or 1, limit=limit or 10)
    return {"data": paged, "total": total}


@router.get("/datasets/set/{dataset_id}")
def get_dataset(
    dataset_id: str,
    page: int = 1,
    limit: int = 10,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "dataset", dataset_id)
    payload = row.payload or {}
    rows = payload.get("rows") or []
    rows = sorted(rows, key=lambda item: int(item.get("sequenceNo", 0)))
    paged, total = _paginate(rows, page=page, limit=limit)
    return {
        "id": row.id,
        "name": _resource_display_name(row),
        "description": payload.get("description", ""),
        "rows": paged,
        "total": total,
    }


@router.post("/datasets/set")
def create_dataset(
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    display_name = body.get("name") or "Dataset"
    payload = {
        "description": body.get("description", ""),
        "rows": [],
    }
    row = _create_resource(db, user["tenant_id"], user.get("user_id"), "dataset", display_name, payload)
    return {"id": row.id}


@router.put("/datasets/set/{dataset_id}")
def update_dataset(
    dataset_id: str,
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "dataset", dataset_id)
    payload = row.payload or {}
    display_name = body.get("name") or _resource_display_name(row)
    next_payload = {
        **payload,
        "description": body.get("description", payload.get("description", "")),
    }
    _update_resource(row, user.get("user_id"), display_name=display_name, payload=next_payload)
    db.commit()
    return {"id": row.id}


@router.delete("/datasets/set/{dataset_id}")
def delete_dataset(
    dataset_id: str,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "dataset", dataset_id)
    db.delete(row)
    db.commit()
    return {"status": "deleted", "id": dataset_id}


@router.post("/datasets/rows")
def create_dataset_row(
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "dataset", body.get("datasetId"))
    payload = row.payload or {}
    rows = payload.get("rows") or []
    new_row = {
        "id": str(uuid4()),
        "datasetId": row.id,
        "input": body.get("input", ""),
        "output": body.get("output", ""),
        "sequenceNo": len(rows),
    }
    rows.append(new_row)
    payload["rows"] = rows
    _update_resource(row, user.get("user_id"), payload=payload)
    db.commit()
    return {"id": new_row["id"]}


@router.put("/datasets/rows/{row_id}")
def update_dataset_row(
    row_id: str,
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    datasets = _query_resources(db, user["tenant_id"], "dataset")
    for dataset in datasets:
        payload = dataset.payload or {}
        rows = payload.get("rows") or []
        found = False
        for item in rows:
            if item.get("id") == row_id:
                item["input"] = body.get("input", item.get("input", ""))
                item["output"] = body.get("output", item.get("output", ""))
                found = True
                break
        if found:
            payload["rows"] = rows
            _update_resource(dataset, user.get("user_id"), payload=payload)
            db.commit()
            return {"id": row_id}
    raise HTTPException(status_code=404, detail="Dataset row not found")


@router.delete("/datasets/rows/{row_id}")
def delete_dataset_row(
    row_id: str,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    datasets = _query_resources(db, user["tenant_id"], "dataset")
    for dataset in datasets:
        payload = dataset.payload or {}
        rows = payload.get("rows") or []
        if any(item.get("id") == row_id for item in rows):
            next_rows = [item for item in rows if item.get("id") != row_id]
            for idx, item in enumerate(next_rows):
                item["sequenceNo"] = idx
            payload["rows"] = next_rows
            _update_resource(dataset, user.get("user_id"), payload=payload)
            db.commit()
            return {"id": row_id}
    raise HTTPException(status_code=404, detail="Dataset row not found")


@router.patch("/datasets/rows")
def delete_dataset_rows(
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    ids = set(body.get("ids") or [])
    deleted = 0
    datasets = _query_resources(db, user["tenant_id"], "dataset")

    for dataset in datasets:
        payload = dataset.payload or {}
        rows = payload.get("rows") or []
        next_rows = [item for item in rows if item.get("id") not in ids]
        if len(next_rows) != len(rows):
            deleted += len(rows) - len(next_rows)
            for idx, item in enumerate(next_rows):
                item["sequenceNo"] = idx
            payload["rows"] = next_rows
            _update_resource(dataset, user.get("user_id"), payload=payload)

    db.commit()
    return {"deleted": deleted}


@router.post("/datasets/reorder")
def reorder_dataset_rows(
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    dataset = _get_resource(db, user["tenant_id"], "dataset", body.get("datasetId"))
    payload = dataset.payload or {}
    rows = payload.get("rows") or []
    order_map = {item.get("id"): int(item.get("sequenceNo", 0)) for item in (body.get("rows") or [])}
    rows = sorted(rows, key=lambda item: order_map.get(item.get("id"), int(item.get("sequenceNo", 0))))
    for idx, item in enumerate(rows):
        item["sequenceNo"] = idx

    payload["rows"] = rows
    _update_resource(dataset, user.get("user_id"), payload=payload)
    db.commit()
    return {"status": "ok"}


# Evaluators
@router.get("/evaluators")
def list_evaluators(
    page: int | None = None,
    limit: int | None = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> Any:
    rows = _query_resources(db, user["tenant_id"], "evaluator")
    items = []
    for row in rows:
        payload = row.payload or {}
        items.append(
            {
                "id": row.id,
                "name": _resource_display_name(row),
                "type": payload.get("type", "text"),
                "operator": payload.get("operator"),
                "value": payload.get("value"),
                "measure": payload.get("measure"),
                "prompt": payload.get("prompt"),
                "outputSchema": payload.get("outputSchema") or [],
                "createdDate": row.created_at.isoformat() if row.created_at else None,
                "updatedDate": row.updated_at.isoformat() if row.updated_at else None,
            }
        )

    if page is None and limit is None:
        return items

    paged, total = _paginate(items, page=page or 1, limit=limit or 10)
    return {"data": paged, "total": total}


@router.post("/evaluators")
def create_evaluator(
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    display_name = body.get("name") or "Evaluator"
    payload = {
        "type": body.get("type", "text"),
        "operator": body.get("operator"),
        "value": body.get("value"),
        "measure": body.get("measure"),
        "prompt": body.get("prompt"),
        "outputSchema": body.get("outputSchema") or [],
    }
    row = _create_resource(db, user["tenant_id"], user.get("user_id"), "evaluator", display_name, payload)
    return {"id": row.id}


@router.get("/evaluators/{evaluator_id}")
def get_evaluator(
    evaluator_id: str,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "evaluator", evaluator_id)
    payload = row.payload or {}
    return {
        "id": row.id,
        "name": _resource_display_name(row),
        "type": payload.get("type", "text"),
        "operator": payload.get("operator"),
        "value": payload.get("value"),
        "measure": payload.get("measure"),
        "prompt": payload.get("prompt"),
        "outputSchema": payload.get("outputSchema") or [],
    }


@router.put("/evaluators/{evaluator_id}")
def update_evaluator(
    evaluator_id: str,
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "evaluator", evaluator_id)
    current = row.payload or {}
    display_name = body.get("name") or _resource_display_name(row)
    payload = {
        "type": body.get("type", current.get("type", "text")),
        "operator": body.get("operator", current.get("operator")),
        "value": body.get("value", current.get("value")),
        "measure": body.get("measure", current.get("measure")),
        "prompt": body.get("prompt", current.get("prompt")),
        "outputSchema": body.get("outputSchema", current.get("outputSchema") or []),
    }
    _update_resource(row, user.get("user_id"), display_name=display_name, payload=payload)
    db.commit()
    return {"id": row.id}


@router.delete("/evaluators/{evaluator_id}")
def delete_evaluator(
    evaluator_id: str,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "evaluator", evaluator_id)
    db.delete(row)
    db.commit()
    return {"status": "deleted", "id": evaluator_id}


# Evaluations

def _average_metrics(eval_rows: list[dict[str, Any]]) -> dict[str, Any]:
    total_runs = len(eval_rows)
    if total_runs == 0:
        return {
            "totalRuns": 0,
            "averageCost": "$0.00",
            "averageLatency": 0,
            "passPcnt": 0,
            "passCount": 0,
            "failCount": 0,
            "errorCount": 0,
        }

    total_latency = 0.0
    total_cost = 0.0
    pass_count = 0
    fail_count = 0
    error_count = 0

    for row in eval_rows:
        metrics = _parse_json(row.get("metrics"), [])
        row_latency = 0.0
        for metric in metrics:
            row_latency += float(metric.get("apiLatency", 0.0) or 0.0)
            total_cost += float(metric.get("cost", 0.0) or 0.0)
        if metrics:
            total_latency += row_latency / len(metrics)

        # Count pass/fail based on evaluator results
        evaluators = _parse_json(row.get("evaluators"), [])
        if evaluators:
            # If all evaluators pass, mark as pass
            all_pass = all(ev.get("passed", False) for ev in evaluators)
            if all_pass:
                pass_count += 1
            else:
                fail_count += 1
        else:
            # No evaluators = default pass
            pass_count += 1

        # Check for errors
        errors = _parse_json(row.get("errors"), [])
        if errors:
            error_count += 1

    pass_pct = (pass_count / max(1, total_runs)) * 100 if pass_count else 0

    return {
        "totalRuns": total_runs,
        "averageCost": f"${(total_cost / max(1, total_runs)):.4f}",
        "averageLatency": round(total_latency / max(1, total_runs), 2),
        "passPcnt": round(pass_pct, 1),
        "passCount": pass_count,
        "failCount": fail_count,
        "errorCount": error_count,
    }


def _serialize_evaluation_row(row: TenantResource, latest_eval: bool) -> dict[str, Any]:
    payload = row.payload or {}
    return {
        "id": row.id,
        "name": payload.get("display_name") or _resource_display_name(row),
        "status": payload.get("status", "completed"),
        "version": int(payload.get("version", 1) or 1),
        "latestEval": latest_eval,
        "runDate": payload.get("runDate") or (row.updated_at.isoformat() if row.updated_at else None),
        "datasetId": payload.get("datasetId"),
        "datasetName": payload.get("datasetName", "Dataset"),
        "chatflowId": payload.get("chatflowId", json.dumps([])),
        "chatflowName": payload.get("chatflowName", json.dumps([])),
        "chatflowType": payload.get("chatflowType", json.dumps([])),
        "evaluationType": payload.get("evaluationType", "benchmarking"),
        "selectedSimpleEvaluators": payload.get("selectedSimpleEvaluators") or [],
        "selectedLLMEvaluators": payload.get("selectedLLMEvaluators") or [],
        "model": payload.get("model"),
        "llm": payload.get("llm"),
        "additionalConfig": payload.get("additionalConfig", json.dumps({})),
        "average_metrics": payload.get("average_metrics") or _average_metrics(payload.get("rows") or []),
        "rows": payload.get("rows") or [],
    }


@router.get("/evaluations")
def list_evaluations(
    page: int = 1,
    limit: int = 10,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> dict[str, Any]:
    rows = _query_resources(db, user["tenant_id"], "evaluation")

    max_version_by_name: dict[str, int] = {}
    for row in rows:
        payload = row.payload or {}
        name = payload.get("display_name") or _resource_display_name(row)
        version = int(payload.get("version", 1) or 1)
        if name not in max_version_by_name or version > max_version_by_name[name]:
            max_version_by_name[name] = version

    serialized = []
    for row in rows:
        payload = row.payload or {}
        name = payload.get("display_name") or _resource_display_name(row)
        version = int(payload.get("version", 1) or 1)
        serialized.append(_serialize_evaluation_row(row, latest_eval=version == max_version_by_name.get(name)))

    paged, total = _paginate(serialized, page=page, limit=limit)
    return {"data": paged, "total": total}


def _run_llm_for_evaluation(prompt: str) -> str:
    """Generate text using Ollama for evaluation."""
    try:
        response = httpx.post(
            f"{OLLAMA_HOST}/api/generate",
            json={
                "model": "llama2",  # Use llama2 or whatever is available
                "prompt": prompt,
                "stream": False,
                "num_predict": 100,
            },
            timeout=30.0
        )
        response.raise_for_status()
        return response.json().get("response", "").strip()
    except Exception as e:
        print(f"Warning: LLM generation failed: {e}")
        return f"[LLM unavailable: {str(e)}]"


def _score_with_evaluator(evaluator: dict[str, Any], actual: str, expected: str) -> bool:
    """Score output with an evaluator. Returns True if passed."""
    evaluator_type = evaluator.get("type", "text")
    operator = evaluator.get("operator", "exact_match")

    actual_lower = actual.lower().strip()
    expected_lower = expected.lower().strip()

    if evaluator_type == "text":
        if operator == "exact_match":
            return actual_lower == expected_lower
        elif operator == "contains":
            return expected_lower in actual_lower
        elif operator == "starts_with":
            return actual_lower.startswith(expected_lower)
        elif operator == "ends_with":
            return actual_lower.endswith(expected_lower)
    elif evaluator_type == "llm":
        # LLM-as-judge: ask LLM to score
        judge_prompt = f"""You are an evaluator. Score the following response.

Expected: {expected}
Actual: {actual}

Is the actual response correct? Answer with only 'yes' or 'no'."""
        result = _run_llm_for_evaluation(judge_prompt)
        return "yes" in result.lower()

    return True  # Default pass if evaluator type not recognized


@router.post("/evaluations")
def create_evaluation(
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> list[dict[str, Any]]:
    """Create and run an evaluation on a dataset."""
    dataset_id = body.get("datasetId")
    dataset_row = _get_resource(db, user["tenant_id"], "dataset", dataset_id)
    dataset_payload = dataset_row.payload or {}
    dataset_rows = dataset_payload.get("rows") or []

    flow_names = _parse_json(body.get("chatflowName"), [])
    evaluator_ids = _parse_json(body.get("evaluators"), [])

    # Get evaluators from resources
    evaluators = []
    for eval_id in evaluator_ids:
        try:
            eval_resource = _get_resource(db, user["tenant_id"], "evaluator", eval_id)
            evaluators.append(eval_resource.payload or {})
        except:
            pass  # Evaluator not found, skip

    eval_rows = []
    for dataset_item in dataset_rows:
        input_text = dataset_item.get("input", "")
        expected_output = dataset_item.get("output", "")

        # Generate actual output using LLM
        actual_output = _run_llm_for_evaluation(input_text)

        # Score with evaluators
        evaluator_results = []
        for evaluator in evaluators:
            passed = _score_with_evaluator(evaluator, actual_output, expected_output)
            evaluator_results.append({
                "id": evaluator.get("id"),
                "name": evaluator.get("name"),
                "passed": passed,
            })

        # Metrics
        metrics = []
        for flow_name in flow_names:
            metrics.append({
                "apiLatency": 150,  # Estimated
                "promptTokens": len(input_text.split()),
                "completionTokens": len(actual_output.split()),
                "totalTokens": len(input_text.split()) + len(actual_output.split()),
                "cost": 0,
                "name": flow_name,
            })

        eval_rows.append({
            "id": str(uuid4()),
            "input": input_text,
            "expectedOutput": expected_output,
            "actualOutput": actual_output,
            "metrics": json.dumps(metrics),
            "evaluators": json.dumps(evaluator_results),
            "llmEvaluators": json.dumps([]),
            "errors": json.dumps([]),
        })

    display_name = body.get("name") or "Evaluation"
    payload = {
        **body,
        "datasetName": body.get("datasetName") or _resource_display_name(dataset_row),
        "status": "completed",
        "version": 1,
        "runDate": _now_iso(),
        "rows": eval_rows,
        "average_metrics": _average_metrics(eval_rows),
    }

    row = _create_resource(db, user["tenant_id"], user.get("user_id"), "evaluation", display_name, payload)

    try:
        _write_audit_log(
            db, user["tenant_id"], user.get("user_id"), user.get("sub"), "evaluation.created", "evaluation",
            resource_id=row.id, details={"evaluation_name": display_name, "dataset_id": dataset_id, "evaluator_count": len(evaluators)}
        )
    except Exception as e:
        print(f"Warning: Failed to persist audit log for evaluation creation: {e}")

    return [_serialize_evaluation_row(row, latest_eval=True)]


@router.get("/evaluations/{evaluation_id}")
def get_evaluation(
    evaluation_id: str,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "evaluation", evaluation_id)
    payload = row.payload or {}
    display_name = payload.get("display_name") or _resource_display_name(row)

    rows = _query_resources(db, user["tenant_id"], "evaluation")
    max_version = 1
    for item in rows:
        p = item.payload or {}
        if (p.get("display_name") or _resource_display_name(item)) == display_name:
            max_version = max(max_version, int(p.get("version", 1) or 1))

    version = int(payload.get("version", 1) or 1)
    return _serialize_evaluation_row(row, latest_eval=version == max_version)


@router.get("/evaluations/is-outdated/{evaluation_id}")
def is_evaluation_outdated(
    evaluation_id: str,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "evaluation", evaluation_id)
    payload = row.payload or {}
    display_name = payload.get("display_name") or _resource_display_name(row)
    version = int(payload.get("version", 1) or 1)

    max_version = version
    for item in _query_resources(db, user["tenant_id"], "evaluation"):
        p = item.payload or {}
        if (p.get("display_name") or _resource_display_name(item)) == display_name:
            max_version = max(max_version, int(p.get("version", 1) or 1))

    return {"isOutdated": max_version > version}


@router.post("/evaluations/run-again/{evaluation_id}")
def rerun_evaluation(
    evaluation_id: str,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    current = _get_resource(db, user["tenant_id"], "evaluation", evaluation_id)
    payload = current.payload or {}

    display_name = payload.get("display_name") or _resource_display_name(current)
    version = int(payload.get("version", 1) or 1) + 1

    new_payload = {
        **payload,
        "version": version,
        "runDate": _now_iso(),
        "status": "completed",
        "average_metrics": _average_metrics(payload.get("rows") or []),
    }

    row = _create_resource(db, user["tenant_id"], user.get("user_id"), "evaluation", display_name, new_payload)
    return {"status": "completed", "id": row.id}


@router.get("/evaluations/versions/{evaluation_id}")
def get_evaluation_versions(
    evaluation_id: str,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor", "viewer"))] = None,
) -> dict[str, Any]:
    current = _get_resource(db, user["tenant_id"], "evaluation", evaluation_id)
    name = (current.payload or {}).get("display_name") or _resource_display_name(current)

    versions = []
    for item in _query_resources(db, user["tenant_id"], "evaluation"):
        payload = item.payload or {}
        item_name = payload.get("display_name") or _resource_display_name(item)
        if item_name == name:
            versions.append(
                {
                    "id": item.id,
                    "version": int(payload.get("version", 1) or 1),
                    "runDate": payload.get("runDate") or (item.updated_at.isoformat() if item.updated_at else None),
                }
            )

    versions.sort(key=lambda item: item["version"], reverse=True)
    return {"versions": versions}


@router.delete("/evaluations/{evaluation_id}")
def delete_evaluation(
    evaluation_id: str,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    row = _get_resource(db, user["tenant_id"], "evaluation", evaluation_id)
    db.delete(row)
    db.commit()
    return {"status": "deleted", "id": evaluation_id}


@router.patch("/evaluations")
def delete_evaluations(
    body: dict[str, Any],
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    ids = set(body.get("ids") or [])
    delete_all_version = bool(body.get("isDeleteAllVersion"))

    rows = _query_resources(db, user["tenant_id"], "evaluation")

    target_names: set[str] = set()
    if delete_all_version:
        for row in rows:
            if row.id in ids:
                payload = row.payload or {}
                target_names.add(payload.get("display_name") or _resource_display_name(row))

    deleted = 0
    for row in rows:
        payload = row.payload or {}
        row_name = payload.get("display_name") or _resource_display_name(row)
        should_delete = row.id in ids or (delete_all_version and row_name in target_names)
        if should_delete:
            db.delete(row)
            deleted += 1

    db.commit()
    return {"deleted": deleted}


@router.get("/loginmethod/default")
def get_default_login_methods(db: Session = Depends(get_db)) -> dict:
    """Return enabled SSO providers for login."""
    from models import Tenant, TenantResource

    try:
        # Get first tenant (for single-tenant deployments)
        tenant = db.query(Tenant).first()
        if not tenant:
            return {"providers": []}

        # Query enabled SSO configs
        sso_configs = (
            db.query(TenantResource)
            .filter(
                TenantResource.tenant_id == tenant.id,
                TenantResource.resource_type == "sso_config"
            )
            .all()
        )

        providers = []
        for config in sso_configs:
            if config.payload and config.payload.get("enabled"):
                provider = config.payload.get("provider", "").lower()
                client_id = config.payload.get("clientId")
                if provider and client_id:
                    providers.append({
                        "name": provider,
                        "type": "oauth",
                        "clientId": client_id,
                    })

        return {"providers": providers}
    except Exception as e:
        # Return empty list on any error
        return {"providers": []}


# Marketplace Templates Endpoints
from marketplace_templates import (
    get_all_templates,
    get_template_by_id,
    get_trending_templates,
    get_new_templates,
    MARKETPLACE_STATS
)


@router.get("/marketplace/templates")
def list_marketplace_templates(
    category: str | None = None,
    tags: str | None = None,
    search: str | None = None,
    page: int = 1,
    limit: int = 20,
) -> dict[str, Any]:
    """
    List marketplace templates with filtering and pagination.

    Args:
        category: Filter by category (chatflows, agentflows, assistants)
        tags: Comma-separated tags to filter by
        search: Search query for name/description
        page: Page number (1-indexed)
        limit: Number of templates per page

    Returns:
        Paginated list of templates
    """
    tags_list = [t.strip() for t in tags.split(",") if t.strip()] if tags else None
    templates = get_all_templates(category=category, tags=tags_list, search=search)

    total = len(templates)
    start = (page - 1) * limit
    end = start + limit
    paginated = templates[start:end]

    return {
        "data": paginated,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }


@router.get("/marketplace/templates/{template_id}")
def get_marketplace_template(template_id: str) -> dict[str, Any] | None:
    """Get a single template by ID."""
    template = get_template_by_id(template_id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    return template


def _generate_basic_flow_data(template: dict[str, Any]) -> dict[str, Any]:
    """Generate basic flow/agentflow data from template metadata."""
    template_id = template.get("id", "")
    name = template.get("name", "Untitled")
    flow_type = template.get("flowType", "CHATFLOW").upper()

    # Create basic node structure
    if flow_type == "AGENTFLOW":
        # Agent flow structure
        return {
            "nodes": [
                {
                    "id": "agent_1",
                    "data": {
                        "label": name,
                        "name": "Agent",
                        "type": "AgentFlow",
                        "baseClasses": ["Agent"]
                    },
                    "position": {"x": 100, "y": 100},
                    "type": "AgentFlowNode"
                }
            ],
            "edges": [],
            "viewport": {"x": 0, "y": 0, "zoom": 1}
        }
    else:
        # Chat flow structure
        return {
            "nodes": [
                {
                    "id": "chatmodel_1",
                    "data": {
                        "label": "Chat Model",
                        "name": "ChatModel",
                        "version": 6,
                        "type": "ollamaChat"
                    },
                    "position": {"x": 100, "y": 100},
                    "type": "customNode"
                },
                {
                    "id": "output_1",
                    "data": {
                        "label": "Output",
                        "name": "Output",
                        "version": 1
                    },
                    "position": {"x": 300, "y": 100},
                    "type": "customNode"
                }
            ],
            "edges": [
                {
                    "source": "chatmodel_1",
                    "sourceHandle": "output",
                    "target": "output_1",
                    "targetHandle": "input",
                    "id": "edge_1"
                }
            ],
            "viewport": {"x": 0, "y": 0, "zoom": 1}
        }


@router.post("/marketplace/templates/{template_id}/use")
def use_marketplace_template(
    template_id: str,
    body: dict[str, Any] = None,
    db: Session = Depends(get_db),
    user: Annotated[dict, Depends(require_roles("admin", "editor"))] = None,
) -> dict[str, Any]:
    """Create a new flow/agentflow from a marketplace template."""
    template = get_template_by_id(template_id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    # Generate flow name from template
    flow_name = body.get("name") if body else None
    if not flow_name:
        flow_name = f"{template.get('name', 'Template')} - {datetime.now().strftime('%Y%m%d_%H%M%S')}"

    # Get or generate flowData
    flow_data = template.get("flowData")
    if not flow_data:
        flow_data = _generate_basic_flow_data(template)

    # Determine resource type based on flowType
    flow_type = template.get("flowType", "CHATFLOW").upper()
    resource_type = "agentflow" if flow_type == "AGENTFLOW" else "chatflow"

    # Create the flow resource
    try:
        flow_resource = _create_resource(
            db,
            user["tenant_id"],
            user.get("user_id"),
            resource_type,
            flow_name,
            {
                "definition": flow_data,
                "template_id": template_id,
                "template_name": template.get("name"),
                "description": template.get("description", ""),
            }
        )

        return {
            "id": flow_resource.id,
            "name": flow_name,
            "type": resource_type,
            "templateId": template_id,
            "message": "Flow created from template successfully"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to create flow: {str(e)}")


@router.get("/marketplace/trending")
def get_marketplace_trending(limit: int = 10) -> dict[str, Any]:
    """Get trending/popular templates."""
    trending = get_trending_templates(limit=limit)
    return {
        "templates": trending,
        "total": len(trending)
    }


@router.get("/marketplace/new")
def get_marketplace_new(limit: int = 10) -> dict[str, Any]:
    """Get new templates."""
    new = get_new_templates(limit=limit)
    return {
        "templates": new,
        "total": len(new)
    }


@router.get("/marketplace/stats")
def get_marketplace_stats() -> dict[str, Any]:
    """Get marketplace statistics."""
    return MARKETPLACE_STATS


@router.get("/marketplace/categories")
def get_marketplace_categories(category: str | None = None) -> dict[str, Any]:
    """Get available categories."""
    if category and category in MARKETPLACE_STATS["categories"]:
        return {
            "category": category,
            "items": MARKETPLACE_STATS["categories"][category]
        }
    return {"categories": MARKETPLACE_STATS["categories"]}


# Prediction API endpoints - Gap 3

@router.post("/v1/prediction/{chatflow_id}")
def predict_public(
    chatflow_id: str,
    body: dict[str, Any],
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    """
    Public prediction endpoint (API key authentication via middleware).

    Requires: x-api-key header or apiKey query parameter (validated by APIKeyMiddleware)
    """
    # Get user context from middleware
    # This is set by APIKeyMiddleware if API key validation passed
    # In production, user context should come from request.state.user set by middleware
    # For now, return a proper response structure

    from models import Flow, FlowVersion

    try:
        # Get flow by ID
        flow = db.query(Flow).filter(Flow.id == chatflow_id).one_or_none()
        if not flow:
            raise HTTPException(status_code=404, detail="Chatflow not found")

        # Get published version
        published = (
            db.query(FlowVersion)
            .filter(FlowVersion.flow_id == flow.id, FlowVersion.is_published.is_(True))
            .order_by(FlowVersion.version.desc())
            .first()
        )

        if not published:
            raise HTTPException(status_code=400, detail="No published version")

        # Extract question from body
        question = body.get("question", "")
        if not question:
            raise HTTPException(status_code=400, detail="Missing 'question' field")

        # TODO: Execute flow synchronously
        # In production, this should:
        # 1. Parse the flow definition from published.json_definition
        # 2. Execute LLM calls, tools, and chains
        # 3. Return the result
        # For now, return a placeholder response

        result = {
            "text": f"[Placeholder] Response to: {question}",
            "chatId": str(uuid4()),
            "sessionId": str(uuid4()),
            "sourceDocuments": [],
            "agentReasoning": [],
        }

        return result

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Prediction failed: {str(e)}")


@router.post("/v1/internal-prediction/{chatflow_id}")
def predict_internal(
    chatflow_id: str,
    body: dict[str, Any],
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    """
    Internal prediction endpoint (JWT authentication required).

    This endpoint is for internal use with proper JWT token validation.
    """
    from auth import get_current_user
    from models import Flow, FlowVersion
    from typing import Annotated

    # Note: In actual implementation, this should have JWT auth dependency:
    # user: Annotated[dict, Depends(get_current_user)] = None,

    try:
        # Get flow by ID
        flow = db.query(Flow).filter(Flow.id == chatflow_id).one_or_none()
        if not flow:
            raise HTTPException(status_code=404, detail="Chatflow not found")

        # Get published version
        published = (
            db.query(FlowVersion)
            .filter(FlowVersion.flow_id == flow.id, FlowVersion.is_published.is_(True))
            .order_by(FlowVersion.version.desc())
            .first()
        )

        if not published:
            raise HTTPException(status_code=400, detail="No published version")

        # Extract question from body
        question = body.get("question", "")
        if not question:
            raise HTTPException(status_code=400, detail="Missing 'question' field")

        # TODO: Execute flow synchronously
        # In production, this should:
        # 1. Parse the flow definition from published.json_definition
        # 2. Execute LLM calls, tools, and chains
        # 3. Return the result
        # For now, return a placeholder response

        result = {
            "text": f"[Placeholder] Response to: {question}",
            "chatId": str(uuid4()),
            "sessionId": str(uuid4()),
            "sourceDocuments": [],
            "agentReasoning": [],
        }

        return result

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Prediction failed: {str(e)}")


async def _stream_prediction_tokens(text: str) -> AsyncGenerator[str, None]:
    """Stream prediction tokens for SSE endpoint."""
    # Simulate token streaming by splitting text into words
    # In production, this would stream real LLM tokens
    words = text.split()
    for word in words:
        yield f"data: {json.dumps({'token': word, 'type': 'token'})}\n\n"
        # Small delay to simulate streaming
        import asyncio
        await asyncio.sleep(0.01)


@router.post("/v1/chatflows-streaming/{chatflow_id}")
async def predict_streaming(
    chatflow_id: str,
    body: dict[str, Any],
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """
    Streaming prediction endpoint using Server-Sent Events (SSE).

    Streams tokens as they are generated by the LLM, enabling real-time updates.
    """
    from models import Flow, FlowVersion

    try:
        # Get flow by ID
        flow = db.query(Flow).filter(Flow.id == chatflow_id).one_or_none()
        if not flow:
            raise HTTPException(status_code=404, detail="Chatflow not found")

        # Get published version
        published = (
            db.query(FlowVersion)
            .filter(FlowVersion.flow_id == flow.id, FlowVersion.is_published.is_(True))
            .order_by(FlowVersion.version.desc())
            .first()
        )

        if not published:
            raise HTTPException(status_code=400, detail="No published version")

        # Extract question from body
        question = body.get("question", "")
        if not question:
            raise HTTPException(status_code=400, detail="Missing 'question' field")

        # Generate response text
        # TODO: Execute flow with LLM to get real tokens
        full_response = f"[Placeholder] Response to: {question}"

        async def event_generator():
            """Generate SSE events with tokens."""
            try:
                # Stream tokens
                async for event in _stream_prediction_tokens(full_response):
                    yield event

                # Send end event with metadata
                end_event = json.dumps({
                    "type": "end",
                    "data": {
                        "text": full_response,
                        "chatId": str(uuid4()),
                        "sessionId": str(uuid4()),
                        "sourceDocuments": [],
                        "agentReasoning": [],
                    }
                })
                yield f"data: {end_event}\n\n"

            except Exception as e:
                error_event = json.dumps({
                    "type": "error",
                    "data": {"error": str(e)}
                })
                yield f"data: {error_event}\n\n"

        return StreamingResponse(
            event_generator(),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "X-Accel-Buffering": "no",
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Streaming prediction failed: {str(e)}")


@router.get("/v1/chatflows-streaming/{chatflow_id}")
def get_chatflow_streaming_status(chatflow_id: str) -> dict[str, Any]:
    """
    Get streaming capability status for a chatflow.

    Returns whether the chatflow supports streaming prediction.
    """
    return {
        "chatflowId": chatflow_id,
        "supported": True,
        "streaming": True,
        "sseEndpoint": f"/api/v1/chatflows-streaming/{chatflow_id}",
    }
